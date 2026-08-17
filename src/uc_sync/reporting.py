"""Detailed XLSX and Databricks-branded standalone HTML reports."""

from __future__ import annotations

import html
import json
from collections import Counter
from dataclasses import asdict, is_dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from uc_sync.inventory import classify_principal


# Excel rejects cells above 32,767 characters, and the Workspace export API
# rejects files above 10 MB. Long definition blobs stay complete in the XLSX up
# to the Excel limit; the browsable HTML keeps a shorter preview.
XLSX_CELL_LIMIT = 32_000
HTML_CELL_LIMIT = 2_000

OBJECT_SECTION_PREFIX = "OBJECT "

ERROR_STATUSES = {
    "ERROR",
    "FAILED",
    "FAILURE",
    "NOT_IMPLEMENTED",
    "DIFFERENT",
    "MISSING_TARGET",
    "EXTRA_TARGET",
}

# MANUAL_ACTION_REQUIRED is tracked separately from hard failures in
# the import comparison report so operators can act on each bucket.
MANUAL_STATUSES = {"MANUAL_ACTION_REQUIRED"}
SUCCESS_STATUSES = {"SUCCESS", "MATCH", "SKIP_EXISTING"}


def classify_import_status(row: Mapping[str, Any]) -> str:
    """Normalize an import row into SUCCESS / FAILURE / MANUAL / OTHER."""

    status = str(
        row.get("status")
        or row.get("import_status")
        or ""
    ).upper()
    if status in SUCCESS_STATUSES or status.startswith("SUCCESS"):
        return "SUCCESS"
    if status in MANUAL_STATUSES:
        return "MANUAL_ACTION_REQUIRED"
    if status in ERROR_STATUSES or status in {"ERROR", "FAILED", "FAILURE"}:
        return "FAILURE"
    if status in {"PENDING", "DRY_RUN", "SKIPPED"}:
        return "PENDING"
    if row.get("error_code") or row.get("error_message"):
        return "FAILURE"
    return status or "OTHER"


def import_comparison_buckets(
    rows: Iterable[Any],
) -> dict[str, list[dict[str, Any]]]:
    """Split import rows into the four operator-facing buckets."""

    details = normalize_rows(rows)
    buckets: dict[str, list[dict[str, Any]]] = {
        "SUCCESS": [],
        "FAILURE": [],
        "MANUAL_ACTION_REQUIRED": [],
        "PENDING": [],
        "OTHER": [],
    }
    for row in details:
        buckets.setdefault(classify_import_status(row), []).append(row)
    return buckets


def import_comparison_summary_rows(
    buckets: Mapping[str, list[dict[str, Any]]],
    *,
    run_id: str = "",
    summary: Mapping[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """Build the high-level Summary sheet for the import comparison report."""

    summary = dict(summary or {})
    success = len(buckets.get("SUCCESS") or [])
    failure = len(buckets.get("FAILURE") or [])
    manual = len(buckets.get("MANUAL_ACTION_REQUIRED") or [])
    pending = len(buckets.get("PENDING") or [])
    other = len(buckets.get("OTHER") or [])
    total = success + failure + manual + pending + other
    rows: list[dict[str, Any]] = [
        {"metric": "run_id", "value": run_id or summary.get("run_id") or ""},
        {"metric": "mode", "value": summary.get("mode") or "IMPORT"},
        {"metric": "dry_run", "value": summary.get("dry_run")},
        {"metric": "total_objects", "value": total},
        {"metric": "success", "value": success},
        {"metric": "failures", "value": failure},
        {"metric": "manual_action_required", "value": manual},
        {"metric": "pending", "value": pending},
        {"metric": "other", "value": other},
        {
            "metric": "outcome",
            "value": (
                "FAILED"
                if failure
                else (
                    "COMPLETED_WITH_WARNINGS"
                    if manual or pending or other
                    else "SUCCESS"
                )
            ),
        },
    ]
    # Per-object-type breakdown for the Summary sheet.
    by_type: dict[str, Counter] = {}
    for bucket, items in buckets.items():
        for row in items:
            object_type = str(row.get("object_type") or "UNKNOWN")
            by_type.setdefault(object_type, Counter())[bucket] += 1
    for object_type in sorted(by_type):
        counts = by_type[object_type]
        rows.append(
            {
                "metric": f"by_type.{object_type}",
                "value": (
                    f"success={counts.get('SUCCESS', 0)}, "
                    f"failures={counts.get('FAILURE', 0)}, "
                    f"manual={counts.get('MANUAL_ACTION_REQUIRED', 0)}, "
                    f"pending={counts.get('PENDING', 0)}"
                ),
            }
        )
    return rows


def import_detail_columns(row: Mapping[str, Any]) -> dict[str, Any]:
    """Prefer the columns operators need when reviewing import outcomes."""

    preferred = [
        "object_type",
        "source_full_name",
        "target_full_name",
        "full_name",
        "status",
        "action",
        "error_code",
        "message",
        "error_message",
        "ddl_path",
        "grants_path",
        "dependency_level",
        "import_order",
        "source_definition_hash",
        "source_object_id",
    ]
    ordered: dict[str, Any] = {}
    for key in preferred:
        if key in row and row.get(key) not in (None, ""):
            ordered[key] = row.get(key)
    for key, value in row.items():
        if key not in ordered:
            ordered[key] = value
    return ordered



def _scalar(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "value"):
        return value.value
    return json.dumps(value, default=str, sort_keys=True)


def normalize_rows(rows: Iterable[Any]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for row in rows:
        if is_dataclass(row):
            value = asdict(row)
        elif hasattr(row, "to_dict"):
            value = row.to_dict()
        elif isinstance(row, Mapping):
            value = dict(row)
        elif isinstance(row, Sequence) and not isinstance(row, (str, bytes)):
            value = {f"column_{idx + 1}": item for idx, item in enumerate(row)}
        else:
            value = {"value": row}
        normalized.append({str(key): _scalar(item) for key, item in value.items()})
    return normalized


def clip(value: Any, limit: int) -> Any:
    if isinstance(value, str) and len(value) > limit:
        return f"{value[:limit]}… [truncated {len(value) - limit} chars]"
    return value


def clip_rows(
    sections: Mapping[str, list[dict[str, Any]]], limit: int
) -> dict[str, list[dict[str, Any]]]:
    return {
        name: [
            {key: clip(value, limit) for key, value in row.items()} for row in rows
        ]
        for name, rows in sections.items()
    }


def is_error(row: Mapping[str, Any]) -> bool:
    status = str(
        row.get("status")
        or row.get("export_status")
        or row.get("import_status")
        or row.get("validation_status")
        or ""
    ).upper()
    return (
        status in ERROR_STATUSES
        or status in MANUAL_STATUSES
        or bool(row.get("error_code") or row.get("error_message"))
    )


def inventory_storage_sheets(
    details: list[dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    """Build inventory-only storage relationship sheets."""

    table_rows = []
    location_rows = []
    credential_rows = []
    for row in details:
        object_type = str(row.get("object_type") or "").upper()
        if object_type in {"TABLE", "EXTERNAL_TABLE"}:
            table_rows.append(
                {
                    "table_name": row.get("full_name") or row.get("name"),
                    "managed_external": (
                        "EXTERNAL"
                        if object_type == "EXTERNAL_TABLE"
                        else "MANAGED"
                    ),
                    "table_type": row.get("table_type"),
                    "data_source_format": row.get("data_source_format"),
                    "table_path": row.get("storage_location"),
                    "external_location_name": row.get(
                        "external_location_name"
                    ),
                    "storage_credential_name": row.get(
                        "storage_credential_name"
                    ),
                }
            )
        elif object_type == "EXTERNAL_LOCATION":
            location_rows.append(
                {
                    "external_location_name": (
                        row.get("external_location_name")
                        or row.get("full_name")
                        or row.get("name")
                    ),
                    "path": row.get("storage_location"),
                    "storage_credential_name": row.get(
                        "storage_credential_name"
                    ),
                }
            )
        elif object_type == "STORAGE_CREDENTIAL":
            credential_rows.append(
                {
                    "external_credential_name": (
                        row.get("storage_credential_name")
                        or row.get("full_name")
                        or row.get("name")
                    ),
                    "credential_type": row.get("credential_type"),
                    "purpose": row.get("credential_purpose"),
                    "connector_id": row.get("access_connector_id"),
                    "user_assigned_managed_identity_id": row.get(
                        "user_assigned_managed_identity_id"
                    ),
                    "permissions": row.get("credential_permissions"),
                }
            )
    return {
        "Table Paths": table_rows,
        "External Locations": location_rows,
        "Storage Credentials": credential_rows,
    }


def _parse_jsonish(value: Any) -> Any:
    if isinstance(value, (list, dict)) or value is None:
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return []
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return value
    return value


def inventory_permission_sheets(
    details: list[dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    """Build inventory sheets for principals and object-level grants."""

    object_permission_rows: list[dict[str, Any]] = []
    principals: dict[str, dict[str, Any]] = {}

    for row in details:
        object_type = str(row.get("object_type") or "")
        object_name = row.get("full_name") or row.get("name")
        grants = _parse_jsonish(row.get("grants")) or []
        if not isinstance(grants, list):
            continue
        for assignment in grants:
            if not isinstance(assignment, dict):
                continue
            principal = str(assignment.get("principal") or "").strip()
            if not principal:
                continue
            privileges = [
                str(item)
                for item in (_parse_jsonish(assignment.get("privileges")) or [])
            ]
            principal_type = str(
                assignment.get("principal_type")
                or classify_principal(principal)
            )
            object_permission_rows.append(
                {
                    "object_type": object_type,
                    "object_name": object_name,
                    "principal": principal,
                    "principal_type": principal_type,
                    "privileges": ", ".join(privileges),
                }
            )
            bucket = principals.setdefault(
                principal,
                {
                    "principal": principal,
                    "principal_type": principal_type,
                    "object_count": 0,
                    "object_types": set(),
                    "privileges": set(),
                },
            )
            bucket["object_count"] += 1
            if object_type:
                bucket["object_types"].add(object_type)
            bucket["privileges"].update(privileges)

    principal_rows = []
    for principal in sorted(principals):
        bucket = principals[principal]
        principal_rows.append(
            {
                "principal": bucket["principal"],
                "principal_type": bucket["principal_type"],
                "object_count": bucket["object_count"],
                "object_types": ", ".join(sorted(bucket["object_types"])),
                "privileges": ", ".join(sorted(bucket["privileges"])),
            }
        )

    return {
        "Principals": principal_rows,
        "Object Permissions": object_permission_rows,
    }


def strip_source_metadata_rows(
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    return [
        {key: value for key, value in row.items() if key != "source_metadata"}
        for row in rows
    ]


def strip_source_metadata_sections(
    sections: Mapping[str, list[dict[str, Any]]],
) -> dict[str, list[dict[str, Any]]]:
    return {
        name: strip_source_metadata_rows(rows) for name, rows in sections.items()
    }


class ReportService:
    """Writes reports locally, then publishes to a UC Volume when possible."""

    def __init__(
        self,
        volume_root: str,
        run_id: str,
        *,
        local_root: str | None = None,
        fs: Any = None,
    ):
        if not volume_root:
            raise ValueError("report_volume_path is required")
        self.run_id = run_id
        self.volume_root = volume_root.rstrip("/")
        self.fs = fs
        # Shared clusters block /tmp and Volume FUSE writes from Python.
        # Stage under /Workspace, then publish to the UC Volume via dbutils.fs.
        default_local = (
            "/Workspace/Users/vivek.ravichandiran@databricks.com/"
            f"UCSync/report_staging/{run_id}"
        )
        base = Path(local_root or default_local)
        self.root = base / "reports"
        self.root.mkdir(parents=True, exist_ok=True)
        self.published_root = f"{self.volume_root}/run_{run_id}/reports"

    def _publish(self, local_path: Path) -> str:
        """Return durable report path.

        Reports are staged under /Workspace because shared clusters block /tmp.
        A Volume copy is preferred because the Workspace export API rejects
        files above 10 MB, so large reports are only retrievable from the
        Volume. Fall back to the staging path when neither copy succeeds.
        """
        relative = local_path.relative_to(self.root).as_posix()
        staged = str(local_path)
        remote = f"{self.published_root}/{relative}"

        remote_path = Path(remote)
        try:
            remote_path.parent.mkdir(parents=True, exist_ok=True)
            remote_path.write_bytes(local_path.read_bytes())
            return remote
        except OSError:
            pass

        if self.fs is not None:
            try:
                self.fs.mkdirs(remote.rsplit("/", 1)[0])
                self.fs.cp(staged, remote, recurse=False)
                return remote
            except Exception:  # noqa: BLE001 - dbutils raises Py4J errors
                return staged

        return staged

    def write_stage(
        self,
        stage: str,
        rows: Iterable[Any],
        summary: Mapping[str, Any],
    ) -> dict[str, str]:
        stage_name = stage.upper()
        details = normalize_rows(rows)
        errors = [row for row in details if is_error(row)]
        grouped: dict[str, list[dict[str, Any]]] = {}
        for row in details:
            group = str(row.get("object_type") or row.get("stage") or "DETAILS")
            grouped.setdefault(group, []).append(row)
        if not grouped:
            grouped = {"DETAILS": []}
        report_sheets = dict(grouped)
        if stage_name == "INVENTORY":
            report_sheets.update(inventory_storage_sheets(details))
            report_sheets.update(inventory_permission_sheets(details))
        stem = stage.lower()
        xlsx_path = self.root / f"{stem}_report.xlsx"
        html_path = self.root / f"{stem}_report.html"
        self._write_xlsx(
            xlsx_path,
            stage_name,
            summary,
            details,
            errors,
            extra_sheets=report_sheets,
        )
        html_path.write_text(
            self._render_html(stage_name, summary, grouped, errors),
            encoding="utf-8",
        )
        lite_details = strip_source_metadata_rows(details)
        lite_errors = [row for row in lite_details if is_error(row)]
        lite_grouped = strip_source_metadata_sections(grouped)
        lite_sheets = strip_source_metadata_sections(report_sheets)
        lite_xlsx = self.root / f"{stem}_report_no_source_metadata.xlsx"
        lite_html = self.root / f"{stem}_report_no_source_metadata.html"
        self._write_xlsx(
            lite_xlsx,
            f"{stage_name} (no source_metadata)",
            summary,
            lite_details,
            lite_errors,
            extra_sheets=lite_sheets,
        )
        lite_html.write_text(
            self._render_html(
                f"{stage_name} (no source_metadata)",
                summary,
                lite_grouped,
                lite_errors,
            ),
            encoding="utf-8",
        )
        return {
            "xlsx": self._publish(xlsx_path),
            "html": self._publish(html_path),
            "xlsx_no_source_metadata": self._publish(lite_xlsx),
            "html_no_source_metadata": self._publish(lite_html),
            "local_xlsx": str(xlsx_path),
            "local_html": str(html_path),
            "local_xlsx_no_source_metadata": str(lite_xlsx),
            "local_html_no_source_metadata": str(lite_html),
        }

    def write_import_comparison(
        self,
        rows: Iterable[Any],
        summary: Mapping[str, Any] | None = None,
    ) -> dict[str, str]:
        """Write the import comparison report (HTML + XLSX).

        Highlights Summary / Success / Failures / Manual Action Required.
        Failures always include full detail columns in the XLSX.
        """

        buckets = import_comparison_buckets(rows)
        success_rows = [import_detail_columns(r) for r in buckets["SUCCESS"]]
        failure_rows = [import_detail_columns(r) for r in buckets["FAILURE"]]
        manual_rows = [
            import_detail_columns(r) for r in buckets["MANUAL_ACTION_REQUIRED"]
        ]
        pending_rows = [import_detail_columns(r) for r in buckets["PENDING"]]
        other_rows = [import_detail_columns(r) for r in buckets["OTHER"]]
        all_rows = (
            success_rows + failure_rows + manual_rows + pending_rows + other_rows
        )
        comparison_summary = {
            **dict(summary or {}),
            "import_success": len(success_rows),
            "import_failures": len(failure_rows),
            "import_manual_action_required": len(manual_rows),
            "import_pending": len(pending_rows),
            "import_total": len(all_rows),
        }
        summary_rows = import_comparison_summary_rows(
            buckets, run_id=self.run_id, summary=comparison_summary
        )
        # Same section-per-tab shape as every other report, so the shared
        # branded template renders the comparison identically.
        sections: dict[str, list[dict[str, Any]]] = {
            "SUCCESS": success_rows,
            "FAILURES": failure_rows,
            "MANUAL ACTION REQUIRED": manual_rows,
        }
        if pending_rows:
            sections["PENDING"] = pending_rows
        if other_rows:
            sections["OTHER"] = other_rows
        xlsx_sections = {"Comparison Summary": summary_rows, **sections}
        for row in all_rows:
            object_type = str(row.get("object_type") or "UNKNOWN")
            section = f"{OBJECT_SECTION_PREFIX}{object_type}"
            xlsx_sections.setdefault(section, []).append(row)

        errors = failure_rows + manual_rows
        xlsx_path = self.root / "import_comparison_report.xlsx"
        html_path = self.root / "import_comparison_report.html"
        self._write_xlsx(
            xlsx_path,
            "IMPORT COMPARISON",
            comparison_summary,
            all_rows,
            errors,
            extra_sheets=xlsx_sections,
        )
        html_path.write_text(
            self._render_html(
                "IMPORT COMPARISON", comparison_summary, sections, errors
            ),
            encoding="utf-8",
        )
        return {
            "xlsx": self._publish(xlsx_path),
            "html": self._publish(html_path),
            "local_xlsx": str(xlsx_path),
            "local_html": str(html_path),
            "success": len(success_rows),
            "failures": len(failure_rows),
            "manual_action_required": len(manual_rows),
            "total": len(all_rows),
        }

    def write_final(
        self,
        summary: Mapping[str, Any],
        stages: Mapping[str, Iterable[Any]],
    ) -> dict[str, str]:
        normalized = {
            name.upper(): normalize_rows(rows) for name, rows in stages.items()
        }
        all_rows = [
            {"stage": stage, **row}
            for stage, rows in normalized.items()
            for row in rows
        ]
        object_sections: dict[str, list[dict[str, Any]]] = {}
        for row in all_rows:
            object_type = str(row.get("object_type") or "").upper()
            if object_type:
                section = f"{OBJECT_SECTION_PREFIX}{object_type}"
                object_sections.setdefault(section, []).append(row)
        report_sections = {**normalized, **object_sections}
        errors = [row for row in all_rows if is_error(row)]
        xlsx_path = self.root / "uc_sync_detailed_report.xlsx"
        html_path = self.root / "uc_sync_summary.html"
        self._write_xlsx(
            xlsx_path,
            "UC SYNC SUMMARY",
            summary,
            all_rows,
            errors,
            extra_sheets=report_sections,
        )
        html_path.write_text(
            self._render_html("UC SYNC SUMMARY", summary, report_sections, errors),
            encoding="utf-8",
        )
        lite_rows = strip_source_metadata_rows(all_rows)
        lite_errors = [row for row in lite_rows if is_error(row)]
        lite_sections = strip_source_metadata_sections(report_sections)
        lite_xlsx = self.root / "uc_sync_detailed_report_no_source_metadata.xlsx"
        lite_html = self.root / "uc_sync_summary_no_source_metadata.html"
        self._write_xlsx(
            lite_xlsx,
            "UC SYNC SUMMARY (no source_metadata)",
            summary,
            lite_rows,
            lite_errors,
            extra_sheets=lite_sections,
        )
        lite_html.write_text(
            self._render_html(
                "UC SYNC SUMMARY (no source_metadata)",
                summary,
                lite_sections,
                lite_errors,
            ),
            encoding="utf-8",
        )
        return {
            "xlsx": self._publish(xlsx_path),
            "html": self._publish(html_path),
            "xlsx_no_source_metadata": self._publish(lite_xlsx),
            "html_no_source_metadata": self._publish(lite_html),
            "local_xlsx": str(xlsx_path),
            "local_html": str(html_path),
            "local_xlsx_no_source_metadata": str(lite_xlsx),
            "local_html_no_source_metadata": str(lite_html),
        }

    def _write_xlsx(
        self,
        path: Path,
        title: str,
        summary: Mapping[str, Any],
        details: list[dict[str, Any]],
        errors: list[dict[str, Any]],
        extra_sheets: Mapping[str, list[dict[str, Any]]] | None = None,
    ) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as exc:  # pragma: no cover - Databricks dependency
            raise RuntimeError(
                "openpyxl is required for XLSX reports; install requirements.txt "
                "on the Databricks Job compute"
            ) from exc

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        summary_sheet.sheet_view.showGridLines = False
        summary_sheet["A1"] = title
        summary_sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
        summary_sheet["A1"].fill = PatternFill("solid", fgColor="0F172A")
        summary_sheet.merge_cells("A1:D1")
        summary_sheet["A2"] = "Run ID"
        summary_sheet["B2"] = self.run_id
        summary_sheet["A3"] = "Generated (UTC)"
        summary_sheet["B3"] = datetime.now(timezone.utc).isoformat()
        row_number = 5
        for key, value in summary.items():
            summary_sheet.cell(row=row_number, column=1, value=str(key))
            summary_sheet.cell(row=row_number, column=2, value=_scalar(value))
            summary_sheet.cell(row=row_number, column=1).font = Font(bold=True)
            row_number += 1
        summary_sheet.column_dimensions["A"].width = 32
        summary_sheet.column_dimensions["B"].width = 80

        def add_table(sheet_name: str, records: list[dict[str, Any]]) -> None:
            safe_name = sheet_name[:31]
            sheet = workbook.create_sheet(safe_name)
            sheet.sheet_view.showGridLines = False
            columns = list(dict.fromkeys(key for row in records for key in row))
            if not columns:
                sheet["A1"] = "No records"
                return
            for column_index, name in enumerate(columns, start=1):
                cell = sheet.cell(row=1, column=column_index, value=name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1E40AF")
                cell.alignment = Alignment(vertical="center")
            for row_index, record in enumerate(records, start=2):
                error_row = is_error(record)
                for column_index, name in enumerate(columns, start=1):
                    cell = sheet.cell(
                        row=row_index,
                        column=column_index,
                        value=clip(_scalar(record.get(name)), XLSX_CELL_LIMIT),
                    )
                    cell.alignment = Alignment(
                        vertical="top", wrap_text=True
                    )
                    if error_row:
                        cell.fill = PatternFill("solid", fgColor="FEE2E2")
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for idx, name in enumerate(columns, start=1):
                values = [str(name)] + [
                    str(record.get(name) or "") for record in records[:500]
                ]
                width = min(max(len(value) for value in values) + 2, 60)
                sheet.column_dimensions[get_column_letter(idx)].width = max(
                    width, 12
                )

        add_table("Details", details)
        add_table("Errors", errors)
        for name, rows in (extra_sheets or {}).items():
            add_table(name.title(), rows)
        workbook.save(path)

    def _render_html(
        self,
        title: str,
        summary: Mapping[str, Any],
        stages: Mapping[str, list[dict[str, Any]]],
        errors: list[dict[str, Any]],
    ) -> str:
        generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        stage_counts = {stage: len(rows) for stage, rows in stages.items()}
        # Per-object-type sections repeat rows already counted by their stage,
        # so status totals only consider stage sections.
        status_counts = Counter(
            str(
                row.get("status")
                or row.get("export_status")
                or row.get("import_status")
                or row.get("validation_status")
                or "UNKNOWN"
            ).upper()
            for name, rows in stages.items()
            if not name.startswith(OBJECT_SECTION_PREFIX)
            for row in rows
        )
        payload = json.dumps(
            clip_rows(stages, HTML_CELL_LIMIT), default=str
        ).replace("</", "<\\/")
        summary_json = json.dumps(dict(summary), default=str).replace("</", "<\\/")
        nav_items = "".join(
            self._nav_item(stage, count) for stage, count in stage_counts.items()
        )
        cards = "".join(
            self._summary_card(stage, count, index)
            for index, (stage, count) in enumerate(stage_counts.items())
        )
        status_pills = "".join(
            f'<span class="stat-pill"><strong>{count}</strong> '
            f"{html.escape(status)}</span>"
            for status, count in sorted(status_counts.items())
        )
        error_box = (
            f'<div class="errors-box"><strong>{len(errors)} errors or manual '
            "actions captured.</strong> Open the Errors tab for details.</div>"
            if errors
            else '<div class="success-box">No errors captured for this run.</div>'
        )
        return f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{html.escape(title)} – {html.escape(self.run_id)}</title>
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
:root{{--bg:#f8fafc;--surface:#fff;--sidebar:#0f172a;--sidebar-hover:#1e293b;
--sidebar-active:#1e40af;--text:#1e293b;--text-muted:#64748b;--border:#e2e8f0;
--radius:10px;--shadow:0 1px 3px rgb(0 0 0/.1)}}
html,body{{height:100%;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',
'Helvetica Neue',Arial,sans-serif;background:var(--bg);color:var(--text);
font-size:14px;line-height:1.5}}.app{{display:flex;height:100vh;overflow:hidden}}
.sidebar{{width:240px;min-width:240px;background:var(--sidebar);color:#e2e8f0;
display:flex;flex-direction:column;overflow-y:auto}}.sidebar-brand{{padding:20px 16px 12px;
border-bottom:1px solid #1e293b}}.sidebar-logo{{display:flex;align-items:center;
margin-bottom:12px}}.sidebar-brand h1{{font-size:13px;font-weight:700;color:#f1f5f9;
letter-spacing:.5px;text-transform:uppercase}}.sidebar-brand p{{font-size:11px;
color:#64748b;margin-top:2px;word-break:break-all}}.ts{{font-size:10px;color:#475569;
margin-top:6px}}nav ul{{list-style:none;padding:8px 0}}.nav-item{{display:flex;
align-items:center;gap:8px;padding:7px 14px;cursor:pointer;border-radius:6px;
margin:1px 6px;font-size:12.5px;color:#94a3b8}}.nav-item:hover{{background:
var(--sidebar-hover);color:#e2e8f0}}.nav-item.active{{background:var(--sidebar-active);
color:#fff;font-weight:600}}.nav-label{{flex:1;white-space:nowrap;overflow:hidden;
text-overflow:ellipsis}}.nav-badge{{font-size:10px;font-weight:700;color:#fff;
padding:1px 6px;border-radius:20px;min-width:20px;text-align:center}}.main{{flex:1;
overflow-y:auto;padding:24px}}.panel{{display:none}}.panel.active{{display:block}}
.summary-header{{margin-bottom:24px}}.summary-header h2{{font-size:22px;font-weight:700}}
.summary-header p{{color:var(--text-muted);margin-top:4px}}.summary-stats{{display:flex;
gap:16px;margin-top:12px;flex-wrap:wrap}}.stat-pill{{background:var(--surface);
border:1px solid var(--border);border-radius:20px;padding:4px 14px;font-size:12px;
color:var(--text-muted)}}.stat-pill strong{{color:var(--text)}}.cards-grid{{display:grid;
grid-template-columns:repeat(auto-fill,minmax(168px,1fr));gap:12px;margin-bottom:24px}}
.card{{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);
padding:16px;cursor:pointer;display:flex;align-items:center;gap:14px;box-shadow:var(--shadow)}}
.card:hover{{transform:translateY(-2px);border-color:#93c5fd}}.card-icon{{font-size:22px;
width:44px;height:44px;border-radius:10px;display:flex;align-items:center;
justify-content:center;background:#dbeafe;color:#1d4ed8}}.card-count{{font-size:24px;
font-weight:800;line-height:1}}.card-label{{font-size:11.5px;color:var(--text-muted);
margin-top:2px}}.errors-box{{background:#fef2f2;border:1px solid #fecaca;
border-radius:var(--radius);padding:12px 16px;color:#b91c1c;margin-top:16px}}
.success-box{{background:#ecfdf5;border:1px solid #a7f3d0;border-radius:var(--radius);
padding:12px 16px;color:#047857;margin-top:16px}}.panel-header{{display:flex;
align-items:center;gap:12px;padding:16px;background:var(--surface);border:1px solid
var(--border);border-radius:var(--radius) var(--radius) 0 0;flex-wrap:wrap}}
.panel-title{{font-size:18px;font-weight:700}}.panel-count{{font-size:12px;
font-weight:700;padding:3px 10px;border-radius:20px;background:#dbeafe;color:#1d4ed8}}
.panel-controls{{margin-left:auto}}.panel-search{{border:1px solid var(--border);
border-radius:6px;padding:6px 12px;width:280px;background:var(--bg)}}.table-wrap{{overflow:auto;
background:var(--surface);border:1px solid var(--border);border-top:none;
max-height:calc(100vh - 190px)}}table{{width:100%;border-collapse:collapse;font-size:13px}}
thead{{position:sticky;top:0;z-index:2}}th{{background:#f1f5f9;color:var(--text-muted);
font-weight:600;padding:10px 14px;text-align:left;white-space:nowrap;border-bottom:2px
solid var(--border)}}td{{padding:9px 14px;border-bottom:1px solid #f1f5f9;
vertical-align:top;max-width:420px;overflow-wrap:anywhere}}tr:hover td{{background:#f8fafc}}
.badge{{display:inline-block;font-size:11px;font-weight:700;padding:2px 7px;
border-radius:4px;white-space:nowrap}}.badge-green{{background:#dcfce7;color:#15803d}}
.badge-red{{background:#fee2e2;color:#b91c1c}}.badge-yellow{{background:#fef9c3;
color:#854d0e}}.badge-gray{{background:#f1f5f9;color:#475569}}.pagination-bar{{display:flex;
justify-content:space-between;align-items:center;padding:10px 16px;background:#fff;
border:1px solid var(--border);border-top:none;border-radius:0 0 var(--radius) var(--radius)}}
.pager-btn{{border:1px solid var(--border);background:#fff;border-radius:6px;padding:5px 12px;
cursor:pointer}}.pager-btn:disabled{{opacity:.35}}@media(max-width:800px){{.sidebar{{width:190px;
min-width:190px}}.main{{padding:16px}}}}
</style></head><body><div class="app"><aside class="sidebar"><div class="sidebar-brand">
<div class="sidebar-logo"><svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 64 64"
width="48" height="48"><g transform="matrix(.75294118,0,0,.75294118,1.0541175,18.823529)">
<path d="M0,24.8V38.5L41.1,60L82.2,38.5V24.8L67.7,17.4L82.2,9.8V-3.9L41.1,-25
L.1,-3.9H0V9.8L14.5,17.4Z" fill="#db1905"/><polygon points="0,60.8 41.1,82.3
82.2,60.8 67.7,53.4 41.1,67.3 14.5,53.4" transform="translate(0,-36)"
fill="#ff5224"/><polygon points="82.2,32.2 41.1,53.7 0,32.2 41.1,11"
transform="translate(0,-36)" fill="#ff5224"/></g></svg></div><h1>{html.escape(title)}</h1>
<p>Run {html.escape(self.run_id)}</p><div class="ts">Generated {generated}</div></div>
<nav><ul><li class="nav-item active" data-tab="summary"><span>⌂</span>
<span class="nav-label">Summary</span></li>{nav_items}
<li class="nav-item" data-tab="errors"><span>!</span><span class="nav-label">Errors</span>
<span class="nav-badge" style="background:#dc2626">{len(errors)}</span></li></ul></nav>
</aside><main class="main"><section class="panel active" id="panel-summary">
<div class="summary-header"><h2>{html.escape(title)}</h2><p>Detailed Unity Catalog
synchronization report</p><div class="summary-stats"><span class="stat-pill">
<strong>{sum(stage_counts.values())}</strong> detailed records</span>{status_pills}
<span class="stat-pill">Snapshot: <strong>{generated}</strong></span></div></div>
<div class="cards-grid">{cards}</div>{error_box}<div id="summary-data"></div>
</section><section class="panel" id="panel-errors"></section>
{''.join(f'<section class="panel" id="panel-{self._slug(stage)}"></section>' for stage in stages)}
</main></div><script>
const DATA={payload};const SUMMARY={summary_json};const PAGE_SIZE=100;const states={{}};
function esc(v){{return String(v??'').replace(/[&<>"']/g,c=>({{'&':'&amp;','<':'&lt;',
'>':'&gt;','"':'&quot;',"'":'&#39;'}}[c]));}}
function slug(v){{return v.toLowerCase().replace(/[^a-z0-9]+/g,'_').replace(/^_|_$/g,'')}}
function statusCell(v){{const s=String(v??'').toUpperCase();const cls=['SUCCESS','MATCH',
'SKIP_EXISTING','CREATE_OR_SKIP'].includes(s)
?'badge-green':(['MANUAL_ACTION_REQUIRED','MANUAL','PENDING','DRY_RUN'].includes(s)
?'badge-yellow':(['ERROR','FAILED','FAILURE','NOT_IMPLEMENTED','DIFFERENT',
'MISSING_TARGET','EXTRA_TARGET'].includes(s)
?'badge-red':(s.includes('WARN')?'badge-yellow':'badge-gray')));return `<span class="badge ${{cls}}">
${{esc(v)}}</span>`}}
function render(name,rows){{const id=slug(name),panel=document.getElementById('panel-'+id);
states[id]=states[id]||{{page:0,query:''}};const state=states[id];const filtered=rows.filter(r=>
JSON.stringify(r).toLowerCase().includes(state.query));const cols=[...new Set(filtered.flatMap(
r=>Object.keys(r)))];const pages=Math.max(1,Math.ceil(filtered.length/PAGE_SIZE));
state.page=Math.min(state.page,pages-1);const start=state.page*PAGE_SIZE;
const pageRows=filtered.slice(start,start+PAGE_SIZE);panel.innerHTML=`<div class="panel-header">
<div class="panel-title">${{esc(name)}}</div><span class="panel-count">${{filtered.length}}</span>
<div class="panel-controls"><input class="panel-search" placeholder="Search all columns..."
value="${{esc(state.query)}}"></div></div><div class="table-wrap"><table><thead><tr>${{
cols.map(c=>`<th>${{esc(c)}}</th>`).join('')}}</tr></thead><tbody>${{pageRows.length?
pageRows.map(r=>`<tr>${{cols.map(c=>`<td>${{c.toLowerCase().includes('status')?
statusCell(r[c]):esc(r[c])}}</td>`).join('')}}</tr>`).join(''):
`<tr><td colspan="${{Math.max(cols.length,1)}}">No records</td></tr>`}}</tbody></table></div>
<div class="pagination-bar"><span>Showing ${{filtered.length?start+1:0}}–${{
Math.min(start+PAGE_SIZE,filtered.length)}} of ${{filtered.length}}</span><div>
<button class="pager-btn prev" ${{state.page===0?'disabled':''}}>Previous</button>
 <strong>${{state.page+1}} / ${{pages}}</strong> <button class="pager-btn next"
${{state.page>=pages-1?'disabled':''}}>Next</button></div></div>`;
panel.querySelector('input').oninput=e=>{{state.query=e.target.value.toLowerCase();
state.page=0;render(name,rows)}};panel.querySelector('.prev').onclick=()=>{{state.page--;render(name,rows)}};
panel.querySelector('.next').onclick=()=>{{state.page++;render(name,rows)}};}}
for(const [name,rows] of Object.entries(DATA))render(name,rows);
render('ERRORS',Object.entries(DATA).filter(([stage])=>
!stage.startsWith('{OBJECT_SECTION_PREFIX}')).flatMap(([stage,rows])=>rows.filter(r=>{{
const s=String(r.status||r.export_status||r.import_status||r.validation_status||'').toUpperCase();
return ['ERROR','FAILED','FAILURE','MANUAL_ACTION_REQUIRED','NOT_IMPLEMENTED','DIFFERENT',
'MISSING_TARGET','EXTRA_TARGET'].includes(s)||
r.error_code||r.error_message}}).map(r=>({{stage,...r}}))));
document.getElementById('summary-data').innerHTML='<pre style="margin-top:16px;background:#fff;'+
'border:1px solid #e2e8f0;border-radius:10px;padding:16px;white-space:pre-wrap">'+
esc(JSON.stringify(SUMMARY,null,2))+'</pre>';
function show(tab){{document.querySelectorAll('.panel').forEach(p=>p.classList.remove('active'));
document.querySelectorAll('.nav-item').forEach(n=>n.classList.toggle('active',n.dataset.tab===tab));
document.getElementById('panel-'+tab).classList.add('active')}}
document.querySelectorAll('.nav-item').forEach(n=>n.onclick=()=>show(n.dataset.tab));
document.querySelectorAll('.card').forEach(c=>c.onclick=()=>show(c.dataset.tab));
</script></body></html>"""

    @staticmethod
    def _slug(value: str) -> str:
        return "_".join(
            part for part in "".join(
                char.lower() if char.isalnum() else " " for char in value
            ).split()
        )

    def _nav_item(self, stage: str, count: int) -> str:
        slug = self._slug(stage)
        icon, color = self._visual(stage)
        return (
            f'<li class="nav-item" data-tab="{slug}"><span>{icon}</span>'
            f'<span class="nav-label">{html.escape(stage.title())}</span>'
            f'<span class="nav-badge" style="background:{color}">{count}</span>'
            "</li>"
        )

    def _summary_card(self, stage: str, count: int, index: int) -> str:
        slug = self._slug(stage)
        icon, color = self._visual(stage)
        return (
            f'<div class="card" data-tab="{slug}"><div class="card-icon" '
            f'style="background:{color}18;color:{color}">{icon}</div>'
            f'<div><div class="card-count" style="color:{color}">{count}</div>'
            f'<div class="card-label">{html.escape(stage.title())}</div></div></div>'
        )

    @staticmethod
    def _visual(name: str) -> tuple[str, str]:
        visuals = {
            "CATALOG": ("C", "#4f46e5"),
            "SCHEMA": ("S", "#7c3aed"),
            "TABLE": ("T", "#0284c7"),
            "EXTERNAL_TABLE": ("ET", "#0369a1"),
            "VIEW": ("V", "#059669"),
            "DYNAMIC_VIEW": ("DV", "#10b981"),
            "METRIC_VIEW": ("μV", "#0891b2"),
            "MATERIALIZED_VIEW": ("MV", "#d97706"),
            "VOLUME": ("VL", "#ea580c"),
            "EXTERNAL_VOLUME": ("EV", "#be185d"),
            "FUNCTION": ("ƒ", "#0f766e"),
            "MODEL": ("M", "#6d28d9"),
            "INVENTORY": ("I", "#4f46e5"),
            "EXPORT": ("E", "#0284c7"),
            "IMPORT": ("I", "#059669"),
            "VALIDATION": ("✓", "#0f766e"),
            "SUCCESS": ("✓", "#15803d"),
            "FAILURES": ("!", "#b91c1c"),
            "MANUAL ACTION REQUIRED": ("⚑", "#b45309"),
            "PENDING": ("⋯", "#64748b"),
        }
        return visuals.get(name.upper(), ("▦", "#1d4ed8"))
