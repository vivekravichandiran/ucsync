"""Clean, operator-facing migration report (Excel) for the governance model.

One workbook per stage: a sheet per object type (storage credentials, external
locations, catalogs, schemas, volumes, functions, tables, views, …) carrying that
type's captured detail plus an import-status column, then governance-detail sheets
(tags, classic masks/row filters, ABAC policies, derived policy→column matches,
grants) so a reviewer can read, per line, exactly which mask/policy/tag/grant is
applied where (design §9). Storage credentials and external locations show an
explicit SKIPPED status when the utility did not create them (create toggle off).
"""

from __future__ import annotations

import io
import json
import re
from pathlib import Path
from typing import Any, Iterable, Optional

# ABAC MATCH COLUMNS predicates. ``has_tag_value('k','v')`` matches a column
# tagged k=v; ``has_tag('k')`` matches any column carrying key k. The ``has_tag(``
# pattern deliberately requires a ``(`` right after, so it never matches inside
# ``has_tag_value(``.
_HAS_TAG_VALUE = re.compile(r"has_tag_value\(\s*'([^']*)'\s*,\s*'([^']*)'\s*\)", re.I)
_HAS_TAG = re.compile(r"has_tag\(\s*'([^']*)'\s*\)", re.I)


def _parse_match_conditions(match_columns: Iterable[Any]) -> list[tuple[str, str, Optional[str]]]:
    """Parse MATCH COLUMNS expressions into ``(kind, tag_key, tag_value)`` tuples.

    ``kind`` is ``"value"`` (key + exact value) or ``"key"`` (key, any value).
    """
    conds: list[tuple[str, str, Optional[str]]] = []
    for expr in match_columns or []:
        s = str(expr)
        for key, val in _HAS_TAG_VALUE.findall(s):
            conds.append(("value", key, val))
        for key in _HAS_TAG.findall(s):
            conds.append(("key", key, None))
    return conds


def _column_condition_match(
    col_tags: dict[str, Any],
    conditions: list[tuple[str, str, Optional[str]]],
) -> Optional[tuple[str, Any]]:
    """Return the ``(tag_key, actual_value)`` of the first satisfied condition."""
    for kind, key, val in conditions:
        if key in col_tags:
            actual = col_tags[key]
            if kind == "key" or str(actual) == str(val):
                return key, actual
    return None


def _table_in_policy_scope(table_full_name: str, on_type: str, on_securable: str) -> bool:
    """Is ``table_full_name`` under the securable the policy is attached ON?"""
    if not on_securable:
        return False
    parts = table_full_name.split(".")
    on_type = on_type.upper()
    if on_type == "CATALOG":
        return parts[0] == on_securable
    if on_type == "SCHEMA":
        return ".".join(parts[:2]) == on_securable
    if on_type == "TABLE":
        return table_full_name == on_securable
    return False


def _import_status_by_name(import_results: Iterable[dict[str, Any]]) -> dict[str, str]:
    status: dict[str, str] = {}
    for r in import_results or []:
        name = str(r.get("target_full_name") or r.get("full_name") or "")
        if name:
            # Prefer a non-SUCCESS status if any phase for the object flagged one.
            prev = status.get(name)
            cur = str(r.get("status") or "")
            if prev in (None, "SUCCESS", "SKIP_EXISTING", "SKIP_CREATE_DISABLED"):
                status[name] = cur
    return status


def _import_index(
    import_results: Iterable[dict[str, Any]],
) -> dict[str, dict[str, str]]:
    """Map securable name → its **creation** import result (status/action/message).

    The engine appends create/structure results before governance phases, so the
    FIRST result seen per name is the creation one — which is what the per-object
    sheets report. Indexed by both target and source name so it resolves whether
    the report reads the migrated (target) or source inventory.
    """
    idx: dict[str, dict[str, str]] = {}
    for r in import_results or []:
        entry = {
            "status": str(r.get("status") or ""),
            "action": str(r.get("action") or ""),
            "message": str(r.get("message") or ""),
        }
        for key in (r.get("target_full_name"), r.get("full_name")):
            key = str(key or "")
            if key and key not in idx:
                idx[key] = entry
    return idx


def _export_index(
    export_results: Iterable[dict[str, Any]],
) -> dict[str, dict[str, str]]:
    """Map securable name → its export (stage 02) result.

    Keyed by both target and source name so it resolves against the migrated
    inventory the export/import reports read (names are never remapped, so the
    two coincide, but resolving either is robust).
    """
    idx: dict[str, dict[str, str]] = {}
    for r in export_results or []:
        entry = {
            "status": str(r.get("status") or ""),
            "error_code": str(r.get("error_code") or ""),
            "error_message": str(r.get("error_message") or ""),
        }
        for key in (r.get("target_full_name"), r.get("full_name")):
            key = str(key or "")
            if key and key not in idx:
                idx[key] = entry
    return idx


def _render_export_status(entry: Optional[dict[str, str]]) -> str:
    """Human-readable export outcome for an object row (stage 02)."""
    if not entry:
        return ""  # no export in this stage (inventory report)
    status = entry["status"]
    msg = entry.get("error_message") or ""
    tail = f": {msg[:200]}" if msg else ""
    if status == "ERROR":
        return f"FAILED{tail}"
    if status == "SUCCESS_WITH_WARNINGS":
        return f"EXPORTED (with warnings{tail})"
    if status == "DRY_RUN":
        return "DRY RUN (validated, not exported)"
    if status == "SUCCESS":
        return "EXPORTED"
    return status or ""


def _render_import_status(entry: Optional[dict[str, str]]) -> str:
    """Human-readable import outcome for an object row.

    Storage credentials / external locations created outside the utility surface
    as ``SKIP_CREATE_DISABLED`` (create toggle off) — rendered as an explicit
    SKIPPED so the reader never mistakes "we didn't touch it" for "it imported".
    """
    if not entry:
        return ""  # no import in this stage (inventory/export reports)
    status, action, msg = entry["status"], entry["action"], entry["message"]
    tail = f": {msg[:200]}" if msg else ""
    if action == "SKIP_CREATE_DISABLED":
        return "SKIPPED — not created by utility (create toggle off; pre-existing)"
    if status == "FAILURE":
        return f"FAILED{tail}"
    if status == "MANUAL_ACTION_REQUIRED":
        return f"MANUAL ACTION REQUIRED{tail}"
    if status == "PENDING" or action == "DRY_RUN":
        return "DRY RUN (validated, not applied)"
    if action == "SKIP_EXISTING":
        return "ALREADY EXISTS (skipped)"
    if action in ("CREATE", "CREATE_OR_SKIP"):
        return "CREATED"
    if status == "SUCCESS":
        return f"SUCCESS ({action})" if action else "SUCCESS"
    return f"{status} ({action})".strip()


# --- Per-object-type sheet column specs -------------------------------------
# Each column is (header, extractor). A trailing "import_status" column is added
# by the renderer for every type.

def _cell(o: dict[str, Any], *keys: str) -> Any:
    """First non-empty value among top-level then ``definition`` for each key."""
    d = o.get("definition") or {}
    for k in keys:
        v = o.get(k)
        if v not in (None, "", [], {}):
            return v
        v = d.get(k)
        if v not in (None, "", [], {}):
            return v
    return ""


def _truncate(v: Any, n: int = 300) -> str:
    s = str(v or "")
    return s if len(s) <= n else s[:n] + "…"


def _col_count(o: dict[str, Any]) -> Any:
    return len((o.get("definition") or {}).get("columns") or []) or ""


def _has_mask(o: dict[str, Any]) -> str:
    return "yes" if (o.get("definition") or {}).get("column_masks") else ""


def _has_row_filter(o: dict[str, Any]) -> str:
    rf = (o.get("definition") or {}).get("row_filter")
    return "yes" if isinstance(rf, dict) and rf.get("function_name") else ""


_VOLUME_COLS = [
    ("volume", lambda o: o["full_name"]),
    ("volume_type", lambda o: _cell(o, "volume_type")),
    ("storage_location", lambda o: _cell(o, "storage_location")),
    ("comment", lambda o: _cell(o, "comment")),
    ("owner", lambda o: o.get("owner") or ""),
]
_TABLE_COLS = [
    ("table", lambda o: o["full_name"]),
    ("table_type", lambda o: _cell(o, "table_type")),
    ("format", lambda o: _cell(o, "data_source_format")),
    ("storage_location", lambda o: _cell(o, "storage_location")),
    ("columns", _col_count),
    ("column_mask", _has_mask),
    ("row_filter", _has_row_filter),
    ("comment", lambda o: _cell(o, "comment")),
    ("owner", lambda o: o.get("owner") or ""),
]
# Identity functions that make a view "dynamic": its SELECT/WHERE masks columns
# or filters rows based on the querying principal. UC has no classic ALTER-applied
# masks/row filters on views (those are table-only) — a view expresses the same
# protection inside its own definition, so surface which markers are present.
_IDENTITY_MARKERS = (
    "current_user(",
    "session_user(",
    "is_member(",
    "is_account_group_member(",
)


def _view_identity(o: dict[str, Any]) -> str:
    d = o.get("definition") or {}
    text = " ".join(
        str(d.get(k) or "") for k in ("view_definition", "view_original_text")
    ).lower()
    return ", ".join(m.rstrip("(") for m in _IDENTITY_MARKERS if m in text)


def _view_functions(o: dict[str, Any]) -> str:
    """Functions the view applies, from UC's authoritative ``view_dependencies``.

    A view can protect a column by calling a UDF inline
    (``SELECT sec.mask_email(email) AS email``). UC records that as a function
    dependency, so we surface it without parsing SQL — this is how a
    function-masked view (as opposed to an identity-based dynamic view) shows up
    in the report.
    """
    d = o.get("definition") or {}
    vd = d.get("view_dependencies") or {}
    deps = vd.get("dependencies") if isinstance(vd, dict) else vd
    names: list[str] = []
    for item in deps or []:
        if not isinstance(item, dict):
            continue
        fn = item.get("function")
        if isinstance(fn, dict):
            name = fn.get("function_full_name") or fn.get("name")
            if name:
                names.append(str(name))
    return ", ".join(sorted(set(names)))


_VIEW_COLS = [
    ("view", lambda o: o["full_name"]),
    ("columns", _col_count),
    ("identity_aware", _view_identity),
    ("functions_applied", _view_functions),
    ("definition", lambda o: _truncate(_cell(o, "view_definition", "view_original_text"))),
    ("comment", lambda o: _cell(o, "comment")),
    ("owner", lambda o: o.get("owner") or ""),
]

# Ordered: creation-dependency order (creds/locations first, governance-bearing
# securables after). Titles are the sheet names.
_TYPE_SHEETS: list[tuple[str, str, list]] = [
    ("STORAGE_CREDENTIAL", "Storage Credentials", [
        ("credential", lambda o: o["full_name"]),
        ("credential_type", lambda o: _cell(o, "credential_type")),
        ("purpose", lambda o: _cell(o, "credential_purpose")),
        ("access_connector_id", lambda o: _cell(o, "access_connector_id")),
        ("managed_identity_id", lambda o: _cell(o, "user_assigned_managed_identity_id")),
        ("read_only", lambda o: _cell(o, "read_only")),
        ("comment", lambda o: _cell(o, "comment")),
        ("owner", lambda o: o.get("owner") or ""),
    ]),
    ("EXTERNAL_LOCATION", "External Locations", [
        ("external_location", lambda o: o["full_name"]),
        ("url", lambda o: _cell(o, "url", "storage_location")),
        ("credential_name", lambda o: _cell(o, "credential_name", "storage_credential_name")),
        ("read_only", lambda o: _cell(o, "read_only")),
        ("comment", lambda o: _cell(o, "comment")),
        ("owner", lambda o: o.get("owner") or ""),
    ]),
    ("CATALOG", "Catalogs", [
        ("catalog", lambda o: o["full_name"]),
        ("catalog_type", lambda o: _cell(o, "catalog_type")),
        ("isolation_mode", lambda o: _cell(o, "isolation_mode")),
        ("storage_root", lambda o: _cell(o, "storage_root")),
        ("comment", lambda o: _cell(o, "comment")),
        ("owner", lambda o: o.get("owner") or ""),
    ]),
    ("SCHEMA", "Schemas", [
        ("schema", lambda o: o["full_name"]),
        ("comment", lambda o: _cell(o, "comment")),
        ("owner", lambda o: o.get("owner") or ""),
    ]),
    ("VOLUME", "Volumes", _VOLUME_COLS),
    ("EXTERNAL_VOLUME", "External Volumes", _VOLUME_COLS),
    ("FUNCTION", "Functions", [
        ("function", lambda o: o["full_name"]),
        ("returns", lambda o: _cell(o, "data_type", "full_data_type")),
        ("deterministic", lambda o: _cell(o, "is_deterministic")),
        ("routine_body", lambda o: _cell(o, "routine_body")),
        ("comment", lambda o: _cell(o, "comment")),
        ("owner", lambda o: o.get("owner") or ""),
    ]),
    ("TABLE", "Tables", _TABLE_COLS),
    ("EXTERNAL_TABLE", "External Tables", _TABLE_COLS),
    ("STREAMING_TABLE", "Streaming Tables", _TABLE_COLS),
    ("VIEW", "Views", _VIEW_COLS),
    ("DYNAMIC_VIEW", "Dynamic Views", _VIEW_COLS),
    ("MATERIALIZED_VIEW", "Materialized Views", _VIEW_COLS),
    ("METRIC_VIEW", "Metric Views", _VIEW_COLS),
]

# Objects the utility inventories but never creates (recreate-by-hand); each gets
# a lean sheet flagged as inventory-only.
_INVENTORY_ONLY = [
    ("MODEL", "Models"),
    ("CONNECTION", "Connections"),
    ("SERVICE_CREDENTIAL", "Service Credentials"),
    ("FOREIGN_CATALOG", "Foreign Catalogs"),
    ("SHARE", "Shares"),
    ("RECIPIENT", "Recipients"),
    ("PROVIDER", "Providers"),
]


def build_report(
    objects: list[dict[str, Any]],
    out_path: str,
    *,
    stage: Optional[str] = None,
    export_results: Optional[list[dict[str, Any]]] = None,
    import_results: Optional[list[dict[str, Any]]] = None,
    run_id: str = "",
) -> str:
    """Write the migration workbook to ``out_path`` (.xlsx). Returns the path.

    Each stage's report becomes the base for the next, so the per-object-type
    sheets carry stage-appropriate status columns:

    * ``INVENTORY`` — no status columns (nothing has happened yet).
    * ``EXPORT``    — an ``export_status`` column.
    * ``IMPORT``    — both ``export_status`` (carried forward from stage 02) and
      ``import_status`` columns.

    ``stage`` is inferred from which results are supplied when not passed
    explicitly (import → IMPORT, export → EXPORT, neither → INVENTORY).
    """

    from openpyxl import Workbook
    from openpyxl.styles import Font

    if stage is None:
        stage = (
            "IMPORT" if import_results is not None
            else "EXPORT" if export_results is not None
            else "INVENTORY"
        )
    stage = str(stage).upper()

    export_idx = _export_index(export_results or [])
    idx = _import_index(import_results or [])

    def _status_headers() -> list[str]:
        headers: list[str] = []
        if stage in ("EXPORT", "IMPORT"):
            headers.append("export_status")
        if stage == "IMPORT":
            headers.append("import_status")
        return headers

    def _status_cells(o: dict[str, Any]) -> list[str]:
        name = o["full_name"]
        tname = str(o.get("target_full_name") or "")
        cells: list[str] = []
        if stage in ("EXPORT", "IMPORT"):
            cells.append(
                _render_export_status(export_idx.get(name) or export_idx.get(tname))
            )
        if stage == "IMPORT":
            cells.append(
                _render_import_status(idx.get(name) or idx.get(tname))
            )
        return cells

    wb = Workbook()

    def _sheet(title: str, headers: list[str]):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        return ws

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["UC Governance Migration report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["run_id", run_id])
    ws.append(["objects", len(objects)])
    counts: dict[str, int] = {}
    for o in objects:
        counts[o["object_type"]] = counts.get(o["object_type"], 0) + 1
    ws.append([])
    ws.append(["object_type", "count"])
    for k in sorted(counts):
        ws.append([k, counts[k]])
    if stage in ("EXPORT", "IMPORT") and export_results:
        st = {}
        for r in export_results:
            st[r.get("status", "")] = st.get(r.get("status", ""), 0) + 1
        ws.append([])
        ws.append(["export_status", "count"])
        for k in sorted(st):
            ws.append([k, st[k]])
    if stage == "IMPORT" and import_results:
        st = {}
        for r in import_results:
            st[r.get("status", "")] = st.get(r.get("status", ""), 0) + 1
        ws.append([])
        ws.append(["import_status", "count"])
        for k in sorted(st):
            ws.append([k, st[k]])

    # One sheet per object type, each carrying that type's captured detail plus
    # an import-status column. Only types actually present get a sheet.
    by_type: dict[str, list[dict[str, Any]]] = {}
    for o in objects:
        by_type.setdefault(o["object_type"], []).append(o)

    status_headers = _status_headers()

    for obj_type, title, cols in _TYPE_SHEETS:
        rows = by_type.get(obj_type)
        if not rows:
            continue
        ws_t = _sheet(title, [h for h, _ in cols] + status_headers)
        for o in sorted(rows, key=lambda x: x["full_name"]):
            ws_t.append([fn(o) for _, fn in cols] + _status_cells(o))

    for obj_type, title in _INVENTORY_ONLY:
        rows = by_type.get(obj_type)
        if not rows:
            continue
        ws_t = _sheet(title, ["object", "comment", "owner", "note"] + status_headers)
        for o in sorted(rows, key=lambda x: x["full_name"]):
            ws_t.append([
                o["full_name"], _cell(o, "comment"), o.get("owner") or "",
                "inventory-only — recreate manually (out of utility scope)",
            ] + _status_cells(o))

    # Catch-all for any present type not explicitly modeled above (never drop an
    # object silently). ABAC policies have their own dedicated sheets below.
    _known = {t for t, _, _ in _TYPE_SHEETS} | {t for t, _ in _INVENTORY_ONLY} | {"ABAC_POLICY"}
    other = [o for o in objects if o["object_type"] not in _known]
    if other:
        ws_t = _sheet("Other Objects", ["object", "type", "comment", "owner"] + status_headers)
        for o in sorted(other, key=lambda x: (x["object_type"], x["full_name"])):
            ws_t.append([
                o["full_name"], o["object_type"], _cell(o, "comment"),
                o.get("owner") or "",
            ] + _status_cells(o))

    # Tags (object + column grain)
    tags = _sheet("Tags", ["object", "level", "column", "key", "value"])
    for o in objects:
        for k, v in (o.get("tags") or {}).items():
            tags.append([o["full_name"], o["object_type"], "", k, v])
        for col, ctags in ((o.get("definition") or {}).get("column_tags") or {}).items():
            for k, v in ctags.items():
                tags.append([o["full_name"], "COLUMN", col, k, v])

    # Column masks & row filters (classic, ALTER-applied). These are table-only in
    # UC — the securable is a table, external table, materialized view, or
    # streaming table (never a plain/dynamic view), so the first column is the
    # general "object", not "table". Views express equivalent protection in their
    # own definition (see the identity_aware column on the view sheets).
    cm = _sheet(
        "Column Masks & Row Filters",
        ["object", "kind", "column", "function", "using/on cols"],
    )
    for o in objects:
        d = o.get("definition") or {}
        for mask in d.get("column_masks") or []:
            cm.append([
                o["full_name"], "CLASSIC MASK", mask.get("column_name"),
                mask.get("function_name"),
                ", ".join(mask.get("using_column_names") or []),
            ])
        rf = d.get("row_filter")
        if isinstance(rf, dict) and rf.get("function_name"):
            cm.append([
                o["full_name"], "CLASSIC ROW FILTER", "",
                rf.get("function_name"),
                ", ".join(rf.get("input_column_names") or []),
            ])

    # ABAC policies
    abac = _sheet(
        "ABAC Policies",
        ["policy", "type", "on_securable", "function", "match_columns",
         "to", "except"],
    )
    for o in objects:
        if o["object_type"] != "ABAC_POLICY":
            continue
        d = o.get("definition") or {}
        abac.append([
            d.get("policy_name"), d.get("policy_type"), d.get("on_securable"),
            d.get("function_name"), "; ".join(d.get("match_columns") or []),
            ", ".join(d.get("to_principals") or []),
            ", ".join(d.get("except_principals") or []),
        ])

    # Policy → matched columns (DERIVED): resolve each ABAC policy's tag rule
    # against the captured column tags, within the securable it is attached ON.
    # UC does not store a policy→column mapping — the effective set is emergent
    # from (policy ON securable) + (MATCH COLUMNS tag rule) + (column tag
    # assignments) — so this sheet reconstructs it as an auditor aid. Matching is
    # on directly-assigned COLUMN tags (has_tag/has_tag_value evaluate column
    # tags); it does not model tag inheritance from parent securables.
    matched = _sheet(
        "Policy Matched Columns",
        ["policy", "type", "on_securable", "function", "match_rule",
         "matched_table", "matched_column", "tag_key", "tag_value"],
    )
    col_tag_index: dict[str, dict[str, Any]] = {}
    for o in objects:
        ctags = (o.get("definition") or {}).get("column_tags") or {}
        if ctags:
            col_tag_index[o["full_name"]] = ctags
    for o in objects:
        if o["object_type"] != "ABAC_POLICY":
            continue
        d = o.get("definition") or {}
        on_type = str(d.get("on_securable_type") or "")
        on_securable = str(d.get("on_securable") or "")
        conditions = _parse_match_conditions(d.get("match_columns") or [])
        rule = "; ".join(d.get("match_columns") or [])
        base = [d.get("policy_name"), d.get("policy_type"), on_securable,
                d.get("function_name"), rule]
        hits = 0
        for table_fn, ctags in col_tag_index.items():
            if not _table_in_policy_scope(table_fn, on_type, on_securable):
                continue
            for col, tags in ctags.items():
                hit = _column_condition_match(tags, conditions)
                if hit:
                    key, value = hit
                    matched.append(base + [table_fn, col, key, value])
                    hits += 1
        if hits == 0:
            note = (
                "(no captured tagged columns match)" if conditions
                else "(match rule not tag-based / unparsed)"
            )
            matched.append(base + ["", "", "", note])

    # Grants (explicit privilege assignments captured at each securable level).
    # These are the grants the utility replays; UC re-establishes inheritance on
    # the target automatically once the parent-level (catalog/schema) grants are
    # replayed, so inherited-only effective privileges are intentionally not
    # listed here (they are not migrated as per-object grants).
    grants = _sheet("Grants", ["object", "level", "principal", "type", "privileges"])
    for o in objects:
        for g in o.get("grants") or []:
            grants.append([
                o["full_name"], o["object_type"], g.get("principal"),
                g.get("principal_type"),
                ", ".join(g.get("privileges") or []),
            ])

    # UC Volumes are mounted via FUSE, which only supports sequential writes.
    # openpyxl.save() writes a ZIP archive and needs a seekable target, so it
    # fails when handed a /Volumes/... path directly. Build the workbook in an
    # in-memory buffer, then flush the finished bytes in one sequential write —
    # which the Volume FUSE mount does support (same pattern the export/bundle
    # writers use to land files on the Volume).
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    buf = io.BytesIO()
    wb.save(buf)
    Path(out_path).write_bytes(buf.getvalue())
    return out_path


def build_report_from_file(inventory_json: str, out_path: str, **kw) -> str:
    objects = json.loads(Path(inventory_json).read_text(encoding="utf-8"))
    return build_report(objects, out_path, **kw)
