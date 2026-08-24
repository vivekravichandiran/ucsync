"""Migration report (Excel) generation test."""
from __future__ import annotations

from uc_sync.report import build_report


def test_build_report_has_governance_sheets(tmp_path):
    objects = [
        {"object_type": "CATALOG", "full_name": "c", "owner": "me",
         "tags": {"class": "INTERNAL"}, "grants": [
             {"principal": "account users", "principal_type": "GROUP",
              "privileges": ["USE_CATALOG"]}]},
        {"object_type": "TABLE", "full_name": "c.s.t", "owner": "me",
         "definition": {"column_masks": [
             {"column_name": "ssn", "function_name": "c.sec.m", "using_column_names": []}],
             "row_filter": {"function_name": "c.sec.rf", "input_column_names": ["dept"]},
             "column_tags": {"ssn": {"pii": "SSN"}}},
         "tags": {}, "grants": []},
        {"object_type": "ABAC_POLICY", "full_name": "c.s.t#policy:p",
         "definition": {"policy_name": "p", "policy_type": "COLUMN_MASK",
                        "on_securable": "c.s.t", "function_name": "c.sec.m",
                        "match_columns": ["has_tag_value('pii','SSN') AS c"],
                        "to_principals": ["account users"],
                        "except_principals": ["svc@x.com"]}},
    ]
    out = tmp_path / "report.xlsx"
    results = [{"target_full_name": "c.s.t", "status": "SUCCESS"}]
    build_report(objects, str(out), import_results=results, run_id="r1")
    assert out.exists()
    from openpyxl import load_workbook
    wb = load_workbook(out)
    # Per-type sheets appear only for types present; governance sheets always do.
    assert set(wb.sheetnames) == {
        "Summary", "Catalogs", "Tables", "Tags", "Column Masks & Row Filters",
        "ABAC Policies", "Policy Matched Columns", "Grants",
    }
    # ABAC sheet carries the policy with its EXCEPT.
    abac_rows = list(wb["ABAC Policies"].iter_rows(values_only=True))
    assert any("svc@x.com" in str(r) for r in abac_rows)
    # Tags sheet has the column tag.
    tag_rows = list(wb["Tags"].iter_rows(values_only=True))
    assert any("SSN" in str(r) for r in tag_rows)
    # The table's per-type sheet carries its import status.
    table_rows = list(wb["Tables"].iter_rows(values_only=True))
    assert any("c.s.t" in str(r) for r in table_rows)


def test_policy_matched_columns_derives_from_tags_within_scope(tmp_path):
    """The derived sheet resolves each policy's tag rule against captured column
    tags, scoped to the securable the policy is attached ON."""
    objects = [
        # Schema-scoped column-mask policy matching pii=SSN.
        {"object_type": "ABAC_POLICY", "full_name": "c.hr#policy:mask_ssn",
         "definition": {"policy_name": "mask_ssn", "policy_type": "COLUMN_MASK",
                        "on_securable_type": "SCHEMA", "on_securable": "c.hr",
                        "function_name": "c.sec.mask", "to_principals": [],
                        "except_principals": [],
                        "match_columns": ["has_tag_value('pii', 'SSN') AS x"]}},
        # In-scope table with a matching column tag and a non-matching one.
        {"object_type": "TABLE", "full_name": "c.hr.employees", "owner": "me",
         "tags": {}, "grants": [],
         "definition": {"column_tags": {"ssn": {"pii": "SSN"},
                                        "email": {"pii": "EMAIL"}}}},
        # Out-of-scope table (different schema) — must NOT match.
        {"object_type": "TABLE", "full_name": "c.finance.ledger", "owner": "me",
         "tags": {}, "grants": [],
         "definition": {"column_tags": {"acct": {"pii": "SSN"}}}},
    ]
    out = tmp_path / "r.xlsx"
    build_report(objects, str(out), run_id="r1")
    from openpyxl import load_workbook
    ws = load_workbook(out)["Policy Matched Columns"]
    rows = list(ws.iter_rows(values_only=True))[1:]  # drop header
    # Exactly one match: c.hr.employees.ssn (pii=SSN), in-scope.
    assert len(rows) == 1
    r = rows[0]
    assert r[0] == "mask_ssn" and r[5] == "c.hr.employees"
    assert r[6] == "ssn" and r[7] == "pii" and r[8] == "SSN"


def test_storage_sheets_status_skipped_vs_created(tmp_path):
    """Storage creds / external locations get their own sheets, and the status
    says SKIPPED when the utility did not create them (create toggle off) vs the
    real import outcome when it did."""
    objects = [
        {"object_type": "STORAGE_CREDENTIAL", "full_name": "cred_a", "owner": "me",
         "credential_type": "AZURE_MANAGED_IDENTITY",
         "access_connector_id": "/subscriptions/x/ac/conn",
         "definition": {"read_only": False}, "grants": []},
        {"object_type": "EXTERNAL_LOCATION", "full_name": "loc_a", "owner": "me",
         "definition": {"url": "abfss://c@acct/p", "credential_name": "cred_a",
                        "read_only": False}, "grants": []},
    ]
    # cred was created by the utility; the external location's create was toggled off.
    results = [
        {"object_type": "STORAGE_CREDENTIAL", "full_name": "cred_a",
         "target_full_name": "cred_a", "status": "SUCCESS", "action": "CREATE"},
        {"object_type": "EXTERNAL_LOCATION", "full_name": "loc_a",
         "target_full_name": "loc_a", "status": "SUCCESS",
         "action": "SKIP_CREATE_DISABLED"},
    ]
    out = tmp_path / "r.xlsx"
    build_report(objects, str(out), import_results=results, run_id="r1")
    from openpyxl import load_workbook
    wb = load_workbook(out)

    cred = list(wb["Storage Credentials"].iter_rows(values_only=True))
    assert cred[0][-1] == "import_status"
    assert cred[1][0] == "cred_a"
    assert "/subscriptions/x/ac/conn" in cred[1]        # access connector captured
    assert cred[1][-1] == "CREATED"

    loc = list(wb["External Locations"].iter_rows(values_only=True))
    assert loc[1][0] == "loc_a"
    assert "abfss://c@acct/p" in loc[1]                  # url captured
    assert "SKIPPED" in loc[1][-1]                        # not created by utility


def test_stage_status_columns(tmp_path):
    """Each stage's report is the base for the next: inventory has no status
    column, export adds export_status, import carries export_status forward and
    adds import_status."""
    from openpyxl import load_workbook

    objects = [{"object_type": "TABLE", "full_name": "c.s.t", "owner": "me",
                "tags": {}, "grants": [], "definition": {}}]

    # INVENTORY — no status columns at all.
    inv = tmp_path / "inventory.xlsx"
    build_report(objects, str(inv), stage="INVENTORY")
    hdr = list(load_workbook(inv)["Tables"].iter_rows(values_only=True))[0]
    assert "export_status" not in hdr and "import_status" not in hdr

    # EXPORT — export_status only, populated from the export results.
    exp = tmp_path / "export.xlsx"
    build_report(objects, str(exp), stage="EXPORT",
                 export_results=[{"full_name": "c.s.t", "status": "SUCCESS"}])
    rows = list(load_workbook(exp)["Tables"].iter_rows(values_only=True))
    assert rows[0][-1] == "export_status" and "import_status" not in rows[0]
    assert rows[1][-1] == "EXPORTED"

    # IMPORT — both columns, export carried forward + this stage's import status.
    imp = tmp_path / "import.xlsx"
    build_report(
        objects, str(imp), stage="IMPORT",
        export_results=[{"full_name": "c.s.t", "status": "SUCCESS"}],
        import_results=[{"target_full_name": "c.s.t", "status": "SUCCESS",
                         "action": "CREATE"}],
    )
    rows = list(load_workbook(imp)["Tables"].iter_rows(values_only=True))
    assert rows[0][-2] == "export_status" and rows[0][-1] == "import_status"
    assert rows[1][-2] == "EXPORTED" and rows[1][-1] == "CREATED"


def test_grants_sheet_has_level_column(tmp_path):
    """The Grants sheet records the securable level of each explicit grant so
    catalog/schema/table/view/function/volume grants are distinguishable."""
    objects = [
        {"object_type": "CATALOG", "full_name": "c", "owner": "me", "tags": {},
         "grants": [{"principal": "u@x.com", "principal_type": "USER",
                     "privileges": ["USE_CATALOG"]}]},
        {"object_type": "TABLE", "full_name": "c.s.t", "owner": "me", "tags": {},
         "grants": [{"principal": "u@x.com", "principal_type": "USER",
                     "privileges": ["SELECT"]}]},
    ]
    out = tmp_path / "r.xlsx"
    build_report(objects, str(out), stage="INVENTORY")
    from openpyxl import load_workbook
    rows = list(load_workbook(out)["Grants"].iter_rows(values_only=True))
    assert rows[0] == ("object", "level", "principal", "type", "privileges")
    levels = {r[0]: r[1] for r in rows[1:]}
    assert levels["c"] == "CATALOG" and levels["c.s.t"] == "TABLE"


def test_masks_sheet_object_header_and_dynamic_view_identity(tmp_path):
    """Classic masks/row filters are table-only, so their sheet's first column is
    'object' (can be an MV/streaming table, never a view). Dynamic views instead
    surface their in-definition protection via an identity_aware column."""
    objects = [
        {"object_type": "TABLE", "full_name": "c.s.t", "owner": "me", "tags": {},
         "grants": [], "definition": {
             "column_masks": [{"column_name": "ssn", "function_name": "c.sec.m",
                               "using_column_names": []}]}},
        {"object_type": "DYNAMIC_VIEW", "full_name": "c.s.dv", "owner": "me",
         "tags": {}, "grants": [], "definition": {
             "view_definition": "SELECT id, CASE WHEN is_account_group_member('hr') "
                                "THEN ssn ELSE '***' END AS ssn FROM c.s.t "
                                "WHERE dept = current_user()"}},
    ]
    out = tmp_path / "r.xlsx"
    build_report(objects, str(out), stage="INVENTORY")
    from openpyxl import load_workbook
    wb = load_workbook(out)

    cm = list(wb["Column Masks & Row Filters"].iter_rows(values_only=True))
    assert cm[0][0] == "object"

    dv = list(wb["Dynamic Views"].iter_rows(values_only=True))
    idx = dv[0].index("identity_aware")
    markers = dv[1][idx]
    assert "is_account_group_member" in markers and "current_user" in markers


def test_view_functions_applied_from_dependencies(tmp_path):
    """A view that masks a column by calling a UDF inline is detected via UC's
    view_dependencies (no SQL parsing) and surfaced in functions_applied."""
    objects = [
        {"object_type": "VIEW", "full_name": "c.s.masked", "owner": "me",
         "tags": {}, "grants": [], "definition": {
             "view_definition": "SELECT id, c.sec.mask_email(email) AS email FROM c.s.t",
             "view_dependencies": {"dependencies": [
                 {"function": {"function_full_name": "c.sec.mask_email"}},
                 {"table": {"table_full_name": "c.s.t"}},
             ]}}},
    ]
    out = tmp_path / "r.xlsx"
    build_report(objects, str(out), stage="INVENTORY")
    from openpyxl import load_workbook
    rows = list(load_workbook(out)["Views"].iter_rows(values_only=True))
    idx = rows[0].index("functions_applied")
    assert rows[1][idx] == "c.sec.mask_email"


def test_build_report_saves_to_buffer_not_path(tmp_path, monkeypatch):
    """UC Volumes FUSE mounts reject the seeks openpyxl needs to write a ZIP to a
    path directly, so the workbook must be built in an in-memory buffer and then
    flushed to the Volume with a single sequential write. Pin that contract:
    Workbook.save must be handed a file-like buffer, never a str/Path target."""
    import io
    from pathlib import Path
    from openpyxl import Workbook

    save_targets = []
    real_save = Workbook.save

    def _spy_save(self, target):  # noqa: ANN001
        save_targets.append(target)
        return real_save(self, target)

    monkeypatch.setattr(Workbook, "save", _spy_save, raising=True)

    out = tmp_path / "reports" / "import.xlsx"
    objects = [{"object_type": "CATALOG", "full_name": "c", "owner": "me",
                "tags": {}, "grants": []}]
    build_report(objects, str(out), run_id="r1")

    assert save_targets, "Workbook.save was never called"
    for target in save_targets:
        assert not isinstance(target, (str, Path)), (
            "workbook saved to a filesystem path (fails on Volume FUSE); "
            "it must be saved to an in-memory buffer then written sequentially"
        )
        assert isinstance(target, io.BytesIO)

    assert out.exists()
    from openpyxl import load_workbook
    assert "Catalogs" in load_workbook(out).sheetnames
