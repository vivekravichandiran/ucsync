"""Governance-aware import: ordering, fail-closed protection, and ABAC.

Covers the plan `plans/governance-accounting-and-failclosed.md`:
  #1 order       — functions execute before tables; views after governance/drop.
  #3 fail-closed — a table that cannot be fully protected never survives, and the
                   FAILURE is what the notebook feeds audit/state (not just report).
  #4 ABAC context— ABAC import issues no USE SCHEMA (no SCHEMA_NOT_FOUND).
  #5 ABAC warehouse — ABAC needs the warehouse executor; without it, fail-closed.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

from uc_sync.audit import stage_audit_row
from uc_sync.package_import import PackageImportEngine
from uc_sync.sync_state import state_row_from_import


# --- a SQL executor that models a real target metastore ---------------------


def _fqn(text: str, prefix: str) -> str:
    """Fully-qualified name that follows ``prefix`` in a statement (unquoted)."""
    m = re.search(prefix + r"\s+([`\w.]+)", text, re.IGNORECASE)
    return m.group(1).replace("`", "") if m else ""


class GovSql:
    """Fake executor with a known-functions set, an allowed-tags set, and a live
    set of existing tables — enough to model fail-closed behavior end to end.

    * ``CREATE TABLE`` with an inline ``MASK f`` where ``f`` is unknown raises
      (ROUTINE_NOT_FOUND); otherwise the table is added to the live set.
    * ``ALTER … SET TAGS`` with a tag key not in ``allowed_tags`` raises.
    * ``CREATE VIEW … FROM t`` raises when ``t`` is not in the live set.
    * ``DROP TABLE`` removes the table; ``DESCRIBE`` proves existence.
    """

    def __init__(self, known_functions=(), allowed_tags=()):
        self.statements: list[str] = []
        self.tables: set[str] = set()
        self.known_functions = set(known_functions)
        self.allowed_tags = set(allowed_tags)

    def execute(self, sql: str):
        self.statements.append(sql)
        text = sql.strip()
        upper = text.upper()
        if upper.startswith("USE ") or upper.startswith("CREATE CATALOG") \
                or upper.startswith("CREATE SCHEMA"):
            return None
        if upper.startswith("CREATE FUNCTION") or upper.startswith("CREATE OR REPLACE FUNCTION"):
            self.known_functions.add(_fqn(text, r"FUNCTION(?:\s+IF\s+NOT\s+EXISTS)?"))
            return None
        if upper.startswith("CREATE TABLE") or upper.startswith("CREATE EXTERNAL TABLE"):
            for m in re.finditer(r"\bMASK\s+([`\w.]+)", text, re.IGNORECASE):
                fn = m.group(1).replace("`", "")
                if fn not in self.known_functions:
                    raise RuntimeError(
                        f"[ROUTINE_NOT_FOUND] The function {fn} cannot be resolved"
                    )
            self.tables.add(_fqn(text, r"TABLE(?:\s+IF\s+NOT\s+EXISTS)?"))
            return None
        if "CREATE" in upper and "VIEW" in upper:
            for m in re.finditer(r"\bFROM\s+([`\w.]+)", text, re.IGNORECASE):
                t = m.group(1).replace("`", "")
                if t not in self.tables:
                    raise RuntimeError(f"[TABLE_OR_VIEW_NOT_FOUND] {t} does not exist")
            return None
        if upper.startswith("ALTER") and "SET TAGS" in upper:
            for key, _ in re.findall(r"'([^']+)'\s*=\s*'([^']*)'", text):
                if key not in self.allowed_tags:
                    raise RuntimeError(f"[UNKNOWN TAG POLICY KEY] tag policy '{key}'")
            return None
        if upper.startswith("DROP TABLE"):
            self.tables.discard(_fqn(text, r"DROP\s+TABLE(?:\s+IF\s+EXISTS)?"))
            return None
        if upper.startswith("DESCRIBE"):
            name = text.split()[-1].replace("`", "").rstrip(";")
            if name not in self.tables:
                raise RuntimeError(f"[TABLE_OR_VIEW_NOT_FOUND] {name}")
            return None
        return None


def _write(root: Path, rel: str, body: str) -> None:
    p = root / rel
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(body, encoding="utf-8")


# --- #1 order ---------------------------------------------------------------


def test_functions_before_tables_and_views_after_governance(tmp_path: Path):
    root = tmp_path / "migrated"
    _write(root, "ddl/CATALOG_c.sql", "CREATE CATALOG `c`;\n")
    _write(root, "ddl/SCHEMA_c__s.sql", "CREATE SCHEMA `c`.`s`;\n")
    _write(root, "ddl/FUNCTION_c__s__mask_ssn.sql",
           "CREATE FUNCTION `c`.`s`.`mask_ssn`(v STRING) RETURNS STRING RETURN '***';\n")
    _write(root, "ddl/TABLE_c__s__t.sql",
           "CREATE TABLE `c`.`s`.`t` (ssn STRING MASK `c`.`s`.`mask_ssn`);\n")
    _write(root, "ddl/VIEW_c__s__v.sql",
           "CREATE VIEW `c`.`s`.`v` AS SELECT * FROM `c`.`s`.`t`;\n")
    _write(root, "tags/TABLE_c__s__t.sql",
           "ALTER TABLE `c`.`s`.`t` SET TAGS ('cls' = 'CONF');\n")
    _write(root, "inventory/objects.json", "[]")

    sql = GovSql(allowed_tags={"cls"})
    results = PackageImportEngine(str(root), sql, dry_run=False).run()

    def idx(pred):
        return next(i for i, s in enumerate(sql.statements) if pred(s.upper()))

    fn_i = idx(lambda u: "FUNCTION" in u and u.startswith("CREATE"))
    tbl_i = idx(lambda u: u.startswith("CREATE TABLE"))
    tag_i = idx(lambda u: "SET TAGS" in u)
    view_i = idx(lambda u: "VIEW" in u and u.startswith("CREATE"))
    # functions before tables; governance (tags) before views.
    assert fn_i < tbl_i < tag_i < view_i
    assert all(r.status == "SUCCESS" for r in results), [
        (r.object_type, r.status, r.message) for r in results
    ]


# --- #3 fail-closed demonstration (the requested negative test) -------------


def _failclosed_bundle(tmp_path: Path) -> Path:
    root = tmp_path / "migrated"
    _write(root, "ddl/CATALOG_c.sql", "CREATE CATALOG `c`;\n")
    _write(root, "ddl/SCHEMA_c__hr.sql", "CREATE SCHEMA `c`.`hr`;\n")
    # (a) inline mask referencing a function in a catalog we are NOT migrating.
    _write(root, "ddl/TABLE_c__hr__t_extern_mask.sql",
           "CREATE TABLE `c`.`hr`.`t_extern_mask` "
           "(x STRING MASK `other_cat`.`sec`.`mask_x`);\n")
    # (b) plain table that will be tagged with a governed tag absent on target.
    _write(root, "ddl/TABLE_c__hr__t_bad_tag.sql",
           "CREATE TABLE `c`.`hr`.`t_bad_tag` (id INT);\n")
    # (c) a view that selects from the bad-tag table (post-governance phase).
    _write(root, "ddl/VIEW_c__hr__v_on_bad_tag.sql",
           "CREATE VIEW `c`.`hr`.`v_on_bad_tag` AS SELECT * FROM `c`.`hr`.`t_bad_tag`;\n")
    _write(root, "tags/TABLE_c__hr__t_bad_tag.sql",
           "ALTER TABLE `c`.`hr`.`t_bad_tag` SET TAGS ('missing_gov_tag' = 'x');\n")
    _write(root, "inventory/objects.json", "[]")
    return root


def test_failclosed_drops_and_propagates(tmp_path: Path):
    root = _failclosed_bundle(tmp_path)
    sql = GovSql(known_functions=set(), allowed_tags={"allowed_only"})
    results = PackageImportEngine(str(root), sql, dry_run=False).run()
    by = {(r.object_type, r.target_full_name): r for r in results}

    extern = by[("TABLE", "c.hr.t_extern_mask")]
    bad = by[("TABLE", "c.hr.t_bad_tag")]
    view = by[("VIEW", "c.hr.v_on_bad_tag")]

    # (a) the external-mask table failed at CREATE and was never created.
    assert extern.status == "FAILURE"
    assert "c.hr.t_extern_mask" not in sql.tables

    # (b) the bad-tag table was created then DROPPED fail-closed.
    assert bad.status == "FAILURE"
    assert bad.error_code == "PROTECTION_FAILED"
    assert any(
        "DROP TABLE IF EXISTS" in s and "t_bad_tag" in s for s in sql.statements
    )
    assert "c.hr.t_bad_tag" not in sql.tables

    # (c) the view on the dropped table failed naturally.
    assert view.status == "FAILURE"

    # The governance (tag) op itself is FAILURE/PROTECTION_FAILED.
    tag_row = next(
        r for r in results if r.policies_path and "t_bad_tag" in r.policies_path
    )
    assert tag_row.status == "FAILURE"
    assert tag_row.error_code == "PROTECTION_FAILED"

    # These FAILUREs are exactly what the notebook hands audit/state (not just the
    # report): audit + state rows carry FAILURE for each.
    for r in (extern, bad, view):
        audit = stage_audit_row(run_id="r1", stage="IMPORT", result=r.to_dict())
        assert audit["status"] == "FAILURE"
        state = state_row_from_import(
            batch_id="b", run_id="r1", result=r.to_dict(), ran_by="me",
            utility_version="1",
        )
        assert state["last_sync_status"] == "FAILURE"


def test_failclosed_preexisting_table_marked_failed_not_dropped(tmp_path: Path):
    """A pre-existing (SKIP_EXISTING) table whose tag fails is NEVER dropped (it may
    hold data — only empty shells this run created are droppable), but its own status
    is flipped to FAILURE: a governance failure is an object failure, so it must not
    read success."""
    root = tmp_path / "migrated"
    _write(root, "ddl/TABLE_c__hr__pre.sql", "CREATE TABLE `c`.`hr`.`pre` (id INT);\n")
    _write(root, "tags/TABLE_c__hr__pre.sql",
           "ALTER TABLE `c`.`hr`.`pre` SET TAGS ('missing' = 'x');\n")
    _write(root, "inventory/objects.json", "[]")

    class PreExistingSql(GovSql):
        def execute(self, sql: str):
            u = sql.strip().upper()
            if u.startswith("CREATE TABLE"):
                # Table already present on target → CREATE reports exists.
                self.statements.append(sql)
                self.tables.add("c.hr.pre")
                raise RuntimeError("[TABLE_OR_VIEW_ALREADY_EXISTS] exists")
            return super().execute(sql)

    sql = PreExistingSql(allowed_tags=set())
    results = PackageImportEngine(str(root), sql, dry_run=False).run()
    table_row = next(r for r in results if r.object_type == "TABLE")
    # Never dropped (has data)...
    assert not any("DROP TABLE" in s for s in sql.statements)
    assert "c.hr.pre" in sql.tables
    # ...but its status reflects the governance failure (not success).
    assert table_row.status == "FAILURE"
    assert table_row.error_code == "PROTECTION_FAILED"


# --- #4 + #5 ABAC context + warehouse enforcement ---------------------------


def _abac_bundle(tmp_path: Path, on_type="CATALOG", on_securable="c",
                 policy_full="c#policy:cat_mask") -> Path:
    root = tmp_path / "migrated"
    _write(root, "ddl/CATALOG_c.sql", "CREATE CATALOG `c`;\n")
    _write(root, "ddl/SCHEMA_c__hr.sql", "CREATE SCHEMA `c`.`hr`;\n")
    _write(root, "ddl/TABLE_c__hr__t.sql", "CREATE TABLE `c`.`hr`.`t` (email STRING);\n")
    # Stem must match export._safe_filename("ABAC_POLICY", policy_full).
    from uc_sync.export import _safe_filename
    stem = _safe_filename("ABAC_POLICY", policy_full)
    _write(root, f"abac/{stem}.sql",
           f"CREATE POLICY `cat_mask` ON {on_type} `{on_securable}` "
           "COLUMN MASK `c`.`hr`.`mask_email` TO `account users` "
           "FOR TABLES MATCH COLUMNS has_tag_value('pii','EMAIL') AS x ON COLUMN x;\n")
    _write(root, "inventory/objects.json", json.dumps([
        {"object_type": "ABAC_POLICY", "full_name": policy_full,
         "definition": {"policy_name": "cat_mask", "on_securable_type": on_type,
                        "on_securable": on_securable}},
    ]))
    return root


class WarehouseSql:
    def __init__(self):
        self.statements: list[str] = []

    def execute(self, sql: str):
        self.statements.append(sql)
        return None


def test_abac_runs_on_warehouse_with_no_use_context(tmp_path: Path):
    root = _abac_bundle(tmp_path)
    main = GovSql()
    wh = WarehouseSql()
    results = PackageImportEngine(
        str(root), main, dry_run=False, abac_sql_executor=wh,
    ).run()

    abac_row = next(r for r in results if r.object_type == "ABAC_POLICY")
    assert abac_row.status == "SUCCESS"
    # CREATE POLICY routed to the warehouse executor, not the Spark one.
    assert any("CREATE POLICY" in s for s in wh.statements)
    assert not any("CREATE POLICY" in s for s in main.statements)
    # #4: no USE SCHEMA `policy` (or any USE) issued for the ABAC phase — the
    # warehouse executor only ran the fully-qualified CREATE POLICY.
    assert all("USE " not in s.upper() for s in wh.statements)
    # The result is keyed by the real policy full name (report/audit can find it).
    assert abac_row.target_full_name == "c#policy:cat_mask"


def test_abac_without_warehouse_fails_fast_and_drops_matched_table(tmp_path: Path):
    # Policy attached ON the table so its scope matches the created table.
    root = _abac_bundle(
        tmp_path, on_type="TABLE", on_securable="c.hr.t",
        policy_full="c.hr.t#policy:tmask",
    )
    main = GovSql()
    results = PackageImportEngine(
        str(root), main, dry_run=False, abac_sql_executor=None,
    ).run()

    abac_row = next(r for r in results if r.object_type == "ABAC_POLICY")
    assert abac_row.status == "FAILURE"
    assert abac_row.error_code == "ABAC_WAREHOUSE_REQUIRED"
    # Fail-closed: the table the policy would have protected is dropped.
    table_row = next(r for r in results if r.object_type == "TABLE")
    assert table_row.status == "FAILURE"
    assert table_row.error_code == "PROTECTION_FAILED"
    assert any("DROP TABLE IF EXISTS" in s and "`t`" in s for s in main.statements)


def test_no_abac_needs_no_warehouse(tmp_path: Path):
    root = tmp_path / "migrated"
    _write(root, "ddl/CATALOG_c.sql", "CREATE CATALOG `c`;\n")
    _write(root, "ddl/TABLE_c__s__t.sql", "CREATE TABLE `c`.`s`.`t` (id INT);\n")
    _write(root, "inventory/objects.json", "[]")
    results = PackageImportEngine(
        str(root), GovSql(), dry_run=False, abac_sql_executor=None,
    ).run()
    assert all(r.status == "SUCCESS" for r in results)


def test_abac_dry_run_needs_no_warehouse(tmp_path: Path):
    root = _abac_bundle(tmp_path)
    results = PackageImportEngine(
        str(root), GovSql(), dry_run=True, abac_sql_executor=None,
    ).run()
    abac_row = next(r for r in results if r.object_type == "ABAC_POLICY")
    assert abac_row.status == "PENDING"


def test_governed_tag_failure_fails_the_catalog_object_without_drop(tmp_path: Path):
    """A governed-tag failure on a non-table securable (here a catalog) marks THAT
    object FAILURE — a governance failure is an object failure, on every type — but
    the catalog is never dropped (destructive; may hold succeeded children)."""
    root = tmp_path / "migrated"
    _write(root, "ddl/CATALOG_c.sql", "CREATE CATALOG `c`;\n")
    _write(root, "ddl/SCHEMA_c__hr.sql", "CREATE SCHEMA `c`.`hr`;\n")
    _write(root, "tags/CATALOG_c.sql",
           "ALTER CATALOG `c` SET TAGS ('missing_gov_tag' = 'x');\n")
    _write(root, "inventory/objects.json", "[]")

    sql = GovSql(allowed_tags={"allowed_only"})  # 'missing_gov_tag' not allowed
    results = PackageImportEngine(str(root), sql, dry_run=False).run()
    cat = next(r for r in results if r.object_type == "CATALOG")
    schema = next(r for r in results if r.object_type == "SCHEMA")

    # The catalog object itself is FAILURE (its governance failed)...
    assert cat.status == "FAILURE"
    assert cat.error_code == "PROTECTION_FAILED"
    # ...but it is never dropped, and unrelated children are untouched.
    assert not any("DROP" in s.upper() for s in sql.statements)
    assert schema.status == "SUCCESS"


# --- P2-B view-on-warehouse routing -----------------------------------------


def test_views_created_on_warehouse_executor_fully_qualified(tmp_path: Path):
    """Plan P2-B: with a warehouse executor supplied, CREATE VIEW (+ existence probe,
    grants) runs on the WAREHOUSE executor — a classic Spark cluster errors on a view
    over a masked table. The view is fully 3-level qualified with NO USE context (the
    warehouse runs each statement statelessly). Tables/functions run on the main
    executor."""
    root = tmp_path / "migrated"
    _write(root, "ddl/CATALOG_c.sql", "CREATE CATALOG `c`;\n")
    _write(root, "ddl/SCHEMA_c__hr.sql", "CREATE SCHEMA `c`.`hr`;\n")
    _write(root, "ddl/TABLE_c__hr__emp.sql", "CREATE TABLE `c`.`hr`.`emp` (id INT);\n")
    _write(root, "ddl/VIEW_c__hr__v.sql",
           "CREATE VIEW `c`.`hr`.`v` AS SELECT * FROM `c`.`hr`.`emp`;\n")
    _write(root, "grants/VIEW_c__hr__v.sql",
           "GRANT SELECT ON VIEW `c`.`hr`.`v` TO `account users`;\n")
    _write(root, "inventory/objects.json", "[]")

    main = GovSql()
    wh = GovSql()
    # The view's base table exists on BOTH executors (same target metastore); model
    # that so the warehouse can resolve the FROM and the DESCRIBE existence probe.
    wh.tables.add("c.hr.emp")
    results = PackageImportEngine(
        str(root), main, dry_run=False, abac_sql_executor=wh,
    ).run()

    assert all(r.status == "SUCCESS" for r in results), [
        (r.object_type, r.status, r.message) for r in results
    ]
    # CREATE VIEW ran on the warehouse, not the main executor, fully 3-level
    # qualified with NO USE context.
    assert any("CREATE" in s.upper() and "VIEW" in s.upper() for s in wh.statements)
    assert not any(
        "CREATE" in s.upper() and "VIEW" in s.upper() for s in main.statements
    )
    assert any("`c`.`hr`.`v`" in s for s in wh.statements)
    assert not any(s.upper().startswith("USE ") for s in wh.statements)
    # The view's ordinary grant ran on the warehouse too.
    assert any("GRANT SELECT" in s for s in wh.statements)
    # The table was created on the main executor (not the warehouse).
    assert any(s.upper().startswith("CREATE TABLE") for s in main.statements)


def test_view_on_dropped_table_still_fails_closed_on_warehouse(tmp_path: Path):
    """A view over a table dropped fail-closed still fails naturally, even when the
    view phase runs on the warehouse executor (TABLE_NOT_FOUND)."""
    root = _failclosed_bundle(tmp_path)
    main = GovSql(known_functions=set(), allowed_tags={"allowed_only"})
    wh = GovSql()  # empty warehouse table set → the dropped table isn't present
    results = PackageImportEngine(
        str(root), main, dry_run=False, abac_sql_executor=wh,
    ).run()
    view = next(r for r in results if r.object_type == "VIEW")
    assert view.status == "FAILURE"
    # The view creation was attempted on the warehouse executor.
    assert any("CREATE" in s.upper() and "VIEW" in s.upper() for s in wh.statements)


# --- full happy path (from-scratch): everything succeeds, no drops -----------


def test_full_happy_path_no_drops(tmp_path: Path):
    """End-to-end from-scratch: catalog → schema → function → table (inline mask,
    function present) → ABAC on warehouse → view → grants + governed tag all
    succeed, nothing is dropped, and the report shows no Issues."""
    from uc_sync.export import _safe_filename
    from uc_sync.report import build_report

    root = tmp_path / "migrated"
    _write(root, "ddl/CATALOG_c.sql", "CREATE CATALOG `c`;\n")
    _write(root, "ddl/SCHEMA_c__hr.sql", "CREATE SCHEMA `c`.`hr`;\n")
    _write(root, "ddl/SCHEMA_c__sec.sql", "CREATE SCHEMA `c`.`sec`;\n")
    _write(root, "ddl/FUNCTION_c__sec__mask_email.sql",
           "CREATE FUNCTION `c`.`sec`.`mask_email`(v STRING) RETURNS STRING RETURN '***';\n")
    _write(root, "ddl/TABLE_c__hr__emp.sql",
           "CREATE TABLE `c`.`hr`.`emp` (email STRING MASK `c`.`sec`.`mask_email`);\n")
    _write(root, "ddl/VIEW_c__hr__v.sql",
           "CREATE VIEW `c`.`hr`.`v` AS SELECT * FROM `c`.`hr`.`emp`;\n")
    _write(root, "grants/TABLE_c__hr__emp.sql",
           "GRANT SELECT ON TABLE `c`.`hr`.`emp` TO `account users`;\n")
    _write(root, "tags/TABLE_c__hr__emp.sql",
           "ALTER TABLE `c`.`hr`.`emp` SET TAGS ('pii' = 'EMAIL');\n")
    policy_full = "c.hr.emp#policy:m"
    stem = _safe_filename("ABAC_POLICY", policy_full)
    _write(root, f"abac/{stem}.sql",
           "CREATE POLICY `m` ON TABLE `c`.`hr`.`emp` COLUMN MASK "
           "`c`.`sec`.`mask_email` TO `account users` FOR TABLES "
           "MATCH COLUMNS has_tag_value('pii','EMAIL') AS x ON COLUMN x;\n")
    inventory = [
        {"object_type": "CATALOG", "full_name": "c", "target_full_name": "c",
         "tags": {}, "grants": []},
        {"object_type": "SCHEMA", "full_name": "c.hr", "target_full_name": "c.hr",
         "tags": {}, "grants": []},
        {"object_type": "SCHEMA", "full_name": "c.sec", "target_full_name": "c.sec",
         "tags": {}, "grants": []},
        {"object_type": "FUNCTION", "full_name": "c.sec.mask_email",
         "target_full_name": "c.sec.mask_email", "tags": {}, "grants": []},
        {"object_type": "TABLE", "full_name": "c.hr.emp",
         "target_full_name": "c.hr.emp", "tags": {"pii": "EMAIL"},
         "grants": [{"principal": "account users", "principal_type": "GROUP",
                     "privileges": ["SELECT"]}]},
        {"object_type": "VIEW", "full_name": "c.hr.v", "target_full_name": "c.hr.v",
         "tags": {}, "grants": []},
        {"object_type": "ABAC_POLICY", "full_name": policy_full,
         "target_full_name": policy_full,
         "definition": {"policy_name": "m", "policy_type": "COLUMN_MASK",
                        "on_securable_type": "TABLE", "on_securable": "c.hr.emp",
                        "function_name": "c.sec.mask_email", "to_principals": [],
                        "except_principals": [], "match_columns": []}},
    ]
    _write(root, "inventory/objects.json", json.dumps(inventory))

    main = GovSql(allowed_tags={"pii"})
    wh = WarehouseSql()
    results = PackageImportEngine(
        str(root), main, dry_run=False, abac_sql_executor=wh,
    ).run()

    # Nothing failed, nothing dropped.
    assert all(r.status in ("SUCCESS", "SKIP_EXISTING") for r in results), [
        (r.object_type, r.status, r.message) for r in results
    ]
    assert not any("DROP TABLE" in s for s in main.statements)
    # ABAC ran on the warehouse; grant + tag applied on the main executor.
    assert any("CREATE POLICY" in s for s in wh.statements)
    assert any("GRANT SELECT" in s for s in main.statements)
    assert any("SET TAGS" in s for s in main.statements)

    # Report: object count matches the inventory, Issues sheet is empty.
    out = tmp_path / "import.xlsx"
    build_report(inventory, str(out), stage="IMPORT", run_id="r1",
                 import_results=[r.to_dict() for r in results])
    from openpyxl import load_workbook
    wb = load_workbook(out)
    issues = list(wb["Issues"].iter_rows(values_only=True))
    assert len(issues) == 1  # header only, no problems
    summary = [c for row in wb["Summary"].iter_rows(values_only=True) for c in row]
    assert len(inventory) in summary  # "objects" == inventory count
