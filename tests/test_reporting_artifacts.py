"""Tests for report publishing and size limits on XLSX/HTML artifacts."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

from uc_sync.reporting import (
    HTML_CELL_LIMIT,
    XLSX_CELL_LIMIT,
    ReportService,
)


def _reporter(tmp_path, **kwargs):
    return ReportService(
        str(tmp_path / "volume"),
        "20260101_000000",
        local_root=str(tmp_path / "staging"),
        **kwargs,
    )


def _html_sections(text):
    payload = re.search(r"const DATA=(.*?);const SUMMARY=", text, re.S)
    return json.loads(payload.group(1))


def test_publishes_reports_into_volume_run_directory(tmp_path):
    reporter = _reporter(tmp_path)

    paths = reporter.write_stage(
        "INVENTORY", [{"object_type": "TABLE", "full_name": "c.s.t"}], {"mode": "SYNC"}
    )

    published_root = f"{tmp_path / 'volume'}/run_20260101_000000/reports"
    assert paths["xlsx"].startswith(published_root)
    assert paths["html"].startswith(published_root)
    for path in (paths["xlsx"], paths["html"], paths["local_xlsx"], paths["local_html"]):
        assert path


def test_falls_back_to_dbutils_when_volume_write_fails(tmp_path):
    class Fs:
        def __init__(self):
            self.copied = []

        def mkdirs(self, path):
            self.copied.append(("mkdirs", path))

        def cp(self, src, dst, recurse=False):
            self.copied.append(("cp", src, dst))

    fs = Fs()
    reporter = _reporter(tmp_path, fs=fs)
    # A file where the Volume directory should be makes mkdir/write fail.
    (tmp_path / "volume").write_text("not a directory", encoding="utf-8")

    paths = reporter.write_stage("INVENTORY", [{"full_name": "c.s.t"}], {})

    assert "run_20260101_000000/reports" in paths["html"]
    assert any(entry[0] == "cp" for entry in fs.copied)


def test_long_values_are_clipped_for_excel_and_html(tmp_path):
    reporter = _reporter(tmp_path)
    blob = "x" * (XLSX_CELL_LIMIT + 5_000)

    paths = reporter.write_stage(
        "INVENTORY",
        [
            {
                "object_type": "DYNAMIC_VIEW",
                "full_name": "c.s.v",
                "status": "SUCCESS",
                "source_metadata": blob,
            }
        ],
        {"mode": "INVENTORY"},
    )

    workbook = load_workbook(paths["local_xlsx"], read_only=True)
    lengths = [
        len(cell)
        for name in workbook.sheetnames
        for row in workbook[name].iter_rows(values_only=True)
        for cell in row
        if isinstance(cell, str)
    ]
    workbook.close()
    assert max(lengths) <= XLSX_CELL_LIMIT + 40
    assert max(lengths) > HTML_CELL_LIMIT

    sections = _html_sections(open(paths["local_html"], encoding="utf-8").read())
    html_row = sections["DYNAMIC_VIEW"][0]
    assert len(html_row["source_metadata"]) <= HTML_CELL_LIMIT + 40
    assert "truncated" in html_row["source_metadata"]
    assert html_row["full_name"] == "c.s.v"


def test_final_html_counts_each_row_once(tmp_path):
    reporter = _reporter(tmp_path)
    rows = [
        {"object_type": "EXTERNAL_TABLE", "full_name": "c.s.e1", "status": "MISSING_TARGET"},
        {"object_type": "TABLE", "full_name": "c.s.t1", "status": "MATCH"},
    ]

    paths = reporter.write_final({"mode": "VALIDATE"}, {"VALIDATION": rows})
    text = open(paths["local_html"], encoding="utf-8").read()

    assert '<span class="nav-badge" style="background:#dc2626">1</span>' in text
    assert "<strong>1</strong> MISSING_TARGET" in text
    assert "<strong>1</strong> MATCH" in text
    # The errors tab is built client-side and must skip drill-down sections.
    assert "!stage.startsWith('OBJECT ')" in text


def test_final_html_stays_small_for_wide_metadata(tmp_path):
    reporter = _reporter(tmp_path)
    rows = [
        {
            "object_type": "TABLE",
            "full_name": f"c.s.t{index}",
            "status": "SUCCESS",
            "source_metadata": "y" * 50_000,
        }
        for index in range(50)
    ]

    paths = reporter.write_final(
        {"mode": "SYNC", "failures": 0},
        {"INVENTORY": rows, "IMPORT": rows},
    )

    size = len(open(paths["local_html"], encoding="utf-8").read())
    assert size < 2_000_000
    sections = _html_sections(open(paths["local_html"], encoding="utf-8").read())
    assert "OBJECT TABLE" in sections
    assert len(sections["INVENTORY"]) == 50


def test_inventory_xlsx_has_storage_relationship_sheets(tmp_path):
    reporter = _reporter(tmp_path)
    rows = [
        {
            "object_type": "TABLE",
            "full_name": "c.s.managed_orders",
            "table_type": "MANAGED",
            "data_source_format": "DELTA",
            "storage_location": "abfss://managed/orders",
            "external_location_name": "__managed_location",
            "storage_credential_name": "__managed_credential",
            "grants": [
                {
                    "principal": "data_engineers",
                    "principal_type": "GROUP",
                    "privileges": ["SELECT"],
                }
            ],
        },
        {
            "object_type": "EXTERNAL_TABLE",
            "full_name": "c.s.external_orders",
            "table_type": "EXTERNAL",
            "data_source_format": "DELTA",
            "storage_location": "abfss://external/orders",
            "external_location_name": "external_data",
            "storage_credential_name": "external_credential",
            "grants": [
                {
                    "principal": "alice@example.com",
                    "principal_type": "USER",
                    "privileges": ["SELECT", "MODIFY"],
                }
            ],
        },
        {
            "object_type": "EXTERNAL_LOCATION",
            "full_name": "external_data",
            "storage_location": "abfss://external",
            "external_location_name": "external_data",
            "storage_credential_name": "external_credential",
            "grants": [
                {
                    "principal": "11111111-2222-3333-4444-555555555555",
                    "principal_type": "SERVICE_PRINCIPAL",
                    "privileges": ["CREATE EXTERNAL TABLE"],
                }
            ],
        },
        {
            "object_type": "STORAGE_CREDENTIAL",
            "full_name": "external_credential",
            "storage_credential_name": "external_credential",
            "credential_type": "AZURE_MANAGED_IDENTITY",
            "credential_purpose": "STORAGE",
            "access_connector_id": "/subscriptions/s/accessConnectors/c",
            "user_assigned_managed_identity_id": "managed-identity-id",
            "credential_permissions": [
                {"principal": "admins", "privileges": ["CREATE EXTERNAL LOCATION"]}
            ],
            "grants": [
                {
                    "principal": "admins",
                    "principal_type": "GROUP",
                    "privileges": ["CREATE EXTERNAL LOCATION"],
                }
            ],
        },
    ]

    paths = reporter.write_stage("INVENTORY", rows, {"inventory": 4})
    workbook = load_workbook(paths["local_xlsx"], read_only=True, data_only=True)

    assert "Table Paths" in workbook.sheetnames
    assert "External Locations" in workbook.sheetnames
    assert "Storage Credentials" in workbook.sheetnames
    assert "Principals" in workbook.sheetnames
    assert "Object Permissions" in workbook.sheetnames

    table_rows = list(workbook["Table Paths"].iter_rows(values_only=True))
    table_headers = list(table_rows[0])
    table_data = [dict(zip(table_headers, row)) for row in table_rows[1:]]
    assert len(table_data) == 2
    external = next(
        row for row in table_data if row["managed_external"] == "EXTERNAL"
    )
    assert external["table_path"] == "abfss://external/orders"
    assert external["storage_credential_name"] == "external_credential"

    location_rows = list(
        workbook["External Locations"].iter_rows(values_only=True)
    )
    assert len(location_rows) == 2

    credential_rows = list(
        workbook["Storage Credentials"].iter_rows(values_only=True)
    )
    credential_headers = list(credential_rows[0])
    credential = dict(zip(credential_headers, credential_rows[1]))
    assert credential["credential_type"] == "AZURE_MANAGED_IDENTITY"
    assert credential["connector_id"].endswith("/accessConnectors/c")
    assert "CREATE EXTERNAL LOCATION" in credential["permissions"]

    principal_rows = list(workbook["Principals"].iter_rows(values_only=True))
    principal_headers = list(principal_rows[0])
    principals = [dict(zip(principal_headers, row)) for row in principal_rows[1:]]
    assert {row["principal"] for row in principals} == {
        "alice@example.com",
        "data_engineers",
        "11111111-2222-3333-4444-555555555555",
        "admins",
    }
    alice = next(row for row in principals if row["principal"] == "alice@example.com")
    assert alice["principal_type"] == "USER"

    permission_rows = list(
        workbook["Object Permissions"].iter_rows(values_only=True)
    )
    permission_headers = list(permission_rows[0])
    permissions = [
        dict(zip(permission_headers, row)) for row in permission_rows[1:]
    ]
    assert len(permissions) == 4
    external_perm = next(
        row for row in permissions if row["object_name"] == "c.s.external_orders"
    )
    assert external_perm["object_type"] == "EXTERNAL_TABLE"
    assert external_perm["principal"] == "alice@example.com"
    assert "MODIFY" in external_perm["privileges"]
    workbook.close()

    lite_workbook = load_workbook(
        paths["local_xlsx_no_source_metadata"], read_only=True, data_only=True
    )
    assert "Principals" in lite_workbook.sheetnames
    assert "Object Permissions" in lite_workbook.sheetnames
    lite_workbook.close()
    assert Path(paths["local_html_no_source_metadata"]).exists()
    lite_sections = _html_sections(
        open(paths["local_html_no_source_metadata"], encoding="utf-8").read()
    )
    assert all(
        "source_metadata" not in row
        for section in lite_sections.values()
        for row in section
    )


def test_stage_reports_emit_no_source_metadata_copies(tmp_path):
    reporter = _reporter(tmp_path)
    paths = reporter.write_stage(
        "EXPORT",
        [
            {
                "object_type": "TABLE",
                "full_name": "c.s.t",
                "status": "SUCCESS",
                "source_metadata": {"secret": "blob"},
            }
        ],
        {"export": 1},
    )
    workbook = load_workbook(
        paths["local_xlsx_no_source_metadata"], read_only=True, data_only=True
    )
    detail_rows = list(workbook["Table"].iter_rows(values_only=True))
    headers = list(detail_rows[0])
    assert "source_metadata" not in headers
    workbook.close()
    sections = _html_sections(
        open(paths["local_html_no_source_metadata"], encoding="utf-8").read()
    )
    assert "source_metadata" not in sections["TABLE"][0]


def test_import_comparison_report_separates_success_failure_manual(tmp_path):
    from uc_sync.reporting import classify_import_status

    reporter = _reporter(tmp_path)
    rows = [
        {
            "object_type": "TABLE",
            "source_full_name": "src.s.t1",
            "target_full_name": "tgt.s.t1",
            "status": "SUCCESS",
            "action": "CREATE_OR_SKIP",
            "message": "ok",
            "ddl_path": "/ddl/t1.sql",
        },
        {
            "object_type": "VIEW",
            "source_full_name": "src.s.v1",
            "target_full_name": "tgt.s.v1",
            "status": "FAILURE",
            "action": "CREATE",
            "error_code": "AnalysisException",
            "message": "SCHEMA_NOT_FOUND",
            "ddl_path": "/ddl/v1.sql",
        },
        {
            "object_type": "STORAGE_CREDENTIAL",
            "source_full_name": "cred",
            "target_full_name": "cred",
            "status": "MANUAL_ACTION_REQUIRED",
            "action": "MANUAL",
            "error_code": "MANUAL_SQL_OBJECT",
            "message": "use REST/API",
            "ddl_path": "/ddl/cred.sql",
        },
    ]
    assert classify_import_status(rows[0]) == "SUCCESS"
    assert classify_import_status(rows[1]) == "FAILURE"
    assert classify_import_status(rows[2]) == "MANUAL_ACTION_REQUIRED"

    paths = reporter.write_import_comparison(
        rows, {"mode": "SYNC", "run_id": "20260101_000000"}
    )
    assert paths["success"] == 1
    assert paths["failures"] == 1
    assert paths["manual_action_required"] == 1
    assert paths["total"] == 3

    workbook = load_workbook(paths["local_xlsx"], read_only=True, data_only=True)
    for sheet in (
        "Summary",
        "Details",
        "Errors",
        "Comparison Summary",
        "Success",
        "Failures",
        "Manual Action Required",
    ):
        assert sheet in workbook.sheetnames, sheet

    failure_rows = list(workbook["Failures"].iter_rows(values_only=True))
    headers = list(failure_rows[0])
    assert "source_full_name" in headers
    assert "error_code" in headers
    assert "message" in headers or "error_message" in headers
    assert "ddl_path" in headers
    assert any("SCHEMA_NOT_FOUND" in str(cell) for cell in failure_rows[1])
    workbook.close()

    html = Path(paths["local_html"]).read_text(encoding="utf-8")
    # Must use the same branded sidebar template as every other report.
    assert 'class="app"' in html and 'class="sidebar"' in html
    assert "IMPORT COMPARISON" in html
    assert 'data-tab="summary"' in html
    assert 'data-tab="errors"' in html
    sections = _html_sections(html)
    assert set(sections) == {"SUCCESS", "FAILURES", "MANUAL ACTION REQUIRED"}
    assert sections["FAILURES"][0]["error_code"] == "AnalysisException"
    assert (
        sections["MANUAL ACTION REQUIRED"][0]["status"] == "MANUAL_ACTION_REQUIRED"
    )
    # FAILURE must be treated as an error status by the shared template JS.
    assert "'FAILURE'" in html


def test_comparison_html_matches_stage_report_shell(tmp_path):
    """The comparison HTML shell must be identical to a normal stage report."""
    reporter = _reporter(tmp_path)
    rows = [{"object_type": "TABLE", "full_name": "c.s.t", "status": "SUCCESS"}]

    stage = reporter.write_stage("IMPORT", rows, {"mode": "IMPORT"})
    comparison = reporter.write_import_comparison(rows, {"mode": "IMPORT"})

    def shell(path):
        text = Path(path).read_text(encoding="utf-8")
        return re.sub(r"const DATA=.*?;const SUMMARY=.*?;const PAGE_SIZE", "", text, flags=re.S)

    stage_shell = shell(stage["local_html"])
    comparison_shell = shell(comparison["local_html"])
    for marker in ("<style>", ".sidebar{", "function render(", "class=\"cards-grid\""):
        assert marker in stage_shell and marker in comparison_shell, marker
    # Same CSS block means the visual format is identical.
    css = lambda t: t[t.index("<style>"): t.index("</style>")]
    assert css(stage_shell) == css(comparison_shell)

