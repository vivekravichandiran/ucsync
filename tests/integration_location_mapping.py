"""Deploy and verify mapped external-location/table import on Databricks."""

from __future__ import annotations

import base64
import configparser
import json
import pathlib
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone

from openpyxl import load_workbook


ROOT = pathlib.Path(__file__).resolve().parents[1]
WORKSPACE_ROOT = (
    "/Workspace/Users/vivek.ravichandiran@databricks.com/UCSync"
)
JOB_ID = 53891349850751
PROFILE = "uc-target"
SOURCE_ROOT = (
    "abfss://unity-catalog-storage@dbstorageisbf2ky3sgcdc.dfs.core.windows.net/"
    "7405609958717235/ucsync/ucsync_local/tables"
)
STAMP = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
TARGET_CATALOG = "ril_sandbox_ucsync_local"
TARGET_EXTERNAL_LOCATION = "classic_stable_target_vk"
TARGET_EXTERNAL_LOCATION_URL = (
    "abfss://unity-catalog-storage@dbstorageisbf2ky3sgcdc.dfs.core.windows.net/"
    "7405609958717235"
)
TARGET_ROOT = (
    "abfss://unity-catalog-storage@dbstorageisbf2ky3sgcdc.dfs.core.windows.net/"
    f"7405609958717235/ucsync/location_mapping/{STAMP}/tables"
)
VOLUME = "/Volumes/classic_stable_target_vk/uc_sync_ops/uc_exports"
CSV_PATH = f"{VOLUME}/location_mapping_{STAMP}.csv"


config = configparser.ConfigParser()
config.read("/Users/vivek.ravichandiran/.databrickscfg")
HOST = config[PROFILE]["host"].rstrip("/")
TOKEN = config[PROFILE]["token"]


def raw(method, path, body=None, content_type="application/json"):
    data = (
        json.dumps(body).encode()
        if body is not None and content_type == "application/json"
        else body
    )
    request = urllib.request.Request(
        HOST + path,
        data=data,
        method=method,
        headers={
            "Authorization": f"Bearer {TOKEN}",
            "Content-Type": content_type,
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=300) as response:
            return response.status, response.read()
    except urllib.error.HTTPError as exc:
        return exc.code, exc.read()


def api(method, path, body=None):
    status, payload = raw(method, path, body)
    return status, json.loads(payload.decode() or "{}")


def deploy(local_path, remote_path, notebook=False):
    content = base64.b64encode((ROOT / local_path).read_bytes()).decode()
    request = {
        "path": remote_path,
        "content": content,
        "overwrite": True,
        "format": "SOURCE" if notebook else "AUTO",
    }
    if notebook:
        request["language"] = "PYTHON"
    status, payload = api("POST", "/api/2.0/workspace/import", request)
    if status != 200:
        raise RuntimeError(f"deploy {local_path}: {status} {payload}")


def run_job(mode):
    params = {
        "execution_mode": "LOCAL",
        "mode": mode,
        "dry_run": "false" if mode == "IMPORT" else "true",
        "catalog_mapping_json": json.dumps(
            {"ril_sandbox": TARGET_CATALOG}
        ),
        "location_mapping_csv_path": CSV_PATH,
        "catalogs": "ril_sandbox",
        "schemas": "ril_sandbox.ucsync_local_01",
        "components": "external_tables",
        "include_parents": "true",
        "exclude_object_types": "",
        "export_volume_path": VOLUME,
        "report_volume_path": VOLUME,
        "audit_table": "classic_stable_target_vk.uc_sync_ops.uc_sync_audit",
    }
    status, response = api(
        "POST",
        "/api/2.1/jobs/run-now",
        {"job_id": JOB_ID, "notebook_params": params},
    )
    if status != 200:
        raise RuntimeError(f"submit {mode}: {status} {response}")
    run_id = response["run_id"]
    while True:
        _, run = api("GET", f"/api/2.1/jobs/runs/get?run_id={run_id}")
        state = run.get("state", {})
        if state.get("life_cycle_state") in {
            "TERMINATED",
            "INTERNAL_ERROR",
            "SKIPPED",
        }:
            break
        time.sleep(15)
    task_run_id = (run.get("tasks") or [{}])[0].get("run_id")
    _, output = api(
        "GET", f"/api/2.1/jobs/runs/get-output?run_id={task_run_id}"
    )
    result = (output.get("notebook_output") or {}).get("result") or ""
    if state.get("result_state") != "SUCCESS":
        raise RuntimeError(
            f"{mode} run {run_id} failed: "
            f"{(output.get('error') or result)[:4000]}"
        )
    return run_id, json.loads(result)


for local, remote, notebook in (
    ("src/uc_sync/components.py", f"{WORKSPACE_ROOT}/src/uc_sync/components.py", False),
    ("src/uc_sync/config.py", f"{WORKSPACE_ROOT}/src/uc_sync/config.py", False),
    ("src/uc_sync/import_engine.py", f"{WORKSPACE_ROOT}/src/uc_sync/import_engine.py", False),
    ("src/uc_sync/inventory.py", f"{WORKSPACE_ROOT}/src/uc_sync/inventory.py", False),
    ("src/uc_sync/location_mapping.py", f"{WORKSPACE_ROOT}/src/uc_sync/location_mapping.py", False),
    ("src/uc_sync/mapping.py", f"{WORKSPACE_ROOT}/src/uc_sync/mapping.py", False),
    ("src/uc_sync/models.py", f"{WORKSPACE_ROOT}/src/uc_sync/models.py", False),
    ("src/uc_sync/validation.py", f"{WORKSPACE_ROOT}/src/uc_sync/validation.py", False),
    ("notebooks/UC_Sync_Main.py", f"{WORKSPACE_ROOT}/notebooks/UC_Sync_Main", True),
):
    deploy(local, remote, notebook)
    print("DEPLOYED", local, flush=True)

csv_payload = (
    "source_external_location,source_location,target_external_location,"
    "target_location,target_credential,target_external_location_url\n"
    f"classic_stable_target_vk,{SOURCE_ROOT},{TARGET_EXTERNAL_LOCATION},"
    f"{TARGET_ROOT},classic_stable_target_vk,{TARGET_EXTERNAL_LOCATION_URL}\n"
).encode()
status, payload = raw(
    "PUT",
    "/api/2.0/fs/files"
    + urllib.parse.quote(CSV_PATH)
    + "?overwrite=true",
    csv_payload,
    "text/csv",
)
if status not in {200, 201, 204}:
    raise RuntimeError(f"CSV upload failed: {status} {payload[:500]}")
print("UPLOADED", CSV_PATH, flush=True)

import_run, import_summary = run_job("IMPORT")
print("IMPORT", import_run, json.dumps(import_summary), flush=True)

status, location = api(
    "GET",
    "/api/2.1/unity-catalog/external-locations/"
    + urllib.parse.quote(TARGET_EXTERNAL_LOCATION),
)
assert status == 200, (status, location)
assert location["url"].rstrip("/") == TARGET_EXTERNAL_LOCATION_URL
assert location["credential_name"] == "classic_stable_target_vk"

table_details = []
for index in range(1, 11):
    full_name = (
        f"{TARGET_CATALOG}.ucsync_local_01.external_table_{index:02d}"
    )
    status, table = api(
        "GET",
        "/api/2.1/unity-catalog/tables/" + urllib.parse.quote(full_name),
    )
    assert status == 200, (full_name, status, table)
    expected = f"{TARGET_ROOT}/external_table_{index:02d}"
    assert table["table_type"] == "EXTERNAL"
    assert table["storage_location"].rstrip("/") == expected
    table_details.append(
        {
            "full_name": full_name,
            "table_type": table["table_type"],
            "storage_location": table["storage_location"],
        }
    )

validate_run, validate_summary = run_job("VALIDATE")
print("VALIDATE", validate_run, json.dumps(validate_summary), flush=True)
assert validate_summary["validated"] == 13
assert validate_summary["failures"] == 0

inventory_xlsx = import_summary["reports"]["inventory"]["xlsx"]
status, workbook_bytes = raw(
    "GET", "/api/2.0/fs/files" + urllib.parse.quote(inventory_xlsx)
)
assert status == 200
report_path = ROOT / "docs" / "location-mapping-inventory.xlsx"
report_path.write_bytes(workbook_bytes)
workbook = load_workbook(report_path, read_only=True, data_only=True)
details = list(workbook["Details"].iter_rows(values_only=True))
headers = [str(value) for value in details[0]]
rows = [dict(zip(headers, row)) for row in details[1:]]
workbook.close()
external_rows = [
    row for row in rows if row.get("object_type") == "EXTERNAL_TABLE"
]
assert len(external_rows) == 10
assert all(row.get("table_type") == "EXTERNAL" for row in external_rows)
assert all(row.get("data_source_format") == "DELTA" for row in external_rows)
assert all(str(row.get("storage_location") or "").startswith(SOURCE_ROOT) for row in external_rows)

result = {
    "status": "SUCCESS",
    "import_run_id": import_run,
    "validation_run_id": validate_run,
    "target_catalog": TARGET_CATALOG,
    "target_external_location": TARGET_EXTERNAL_LOCATION,
    "target_root": TARGET_ROOT,
    "location_mapping_csv": CSV_PATH,
    "import_summary": {
        key: import_summary.get(key)
        for key in ("inventory", "imported", "failures", "status")
    },
    "validation_summary": {
        key: validate_summary.get(key)
        for key in ("inventory", "validated", "failures", "status")
    },
    "tables": table_details,
    "inventory_report": str(report_path),
}
(ROOT / "docs" / "location-mapping-integration-result.json").write_text(
    json.dumps(result, indent=2),
    encoding="utf-8",
)
print("LOCATION_MAPPING_INTEGRATION_COMPLETE", json.dumps(result), flush=True)
